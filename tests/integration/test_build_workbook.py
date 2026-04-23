"""Integration tests for scripts/build_workbook.py — pure file I/O, no browser."""

import csv
import shutil
import tempfile
import unittest
from datetime import date
from pathlib import Path

import openpyxl

from tests._helpers import CSV_FIXTURES, ROOT  # noqa: F401

import build_workbook


class TestBuildWorkbook(unittest.TestCase):
    def setUp(self):
        self._tmp     = tempfile.TemporaryDirectory()
        self.work_dir = Path(self._tmp.name)

    def tearDown(self):
        self._tmp.cleanup()

    def test_new_workbook_from_csv(self):
        # Copy the test price CSV (has both a confirmed and In Progress rows)
        shutil.copy(
            CSV_FIXTURES / "copart_price_with_in_progress.csv",
            self.work_dir / "copart_price_2026_01_02.csv",
        )
        workbook_path = self.work_dir / "auction_results.xlsx"

        pending = build_workbook.find_pending_files(
            self.work_dir, date(2026, 6, 1), processed=set(),
        )
        self.assertEqual(len(pending), 1)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        added = build_workbook.process_csv(pending[0], wb)
        self.assertEqual(added, 3)
        wb.save(workbook_path)

        # Verify: one sheet per Make
        wb = openpyxl.load_workbook(workbook_path)
        self.assertEqual(set(wb.sheetnames), {"HONDA", "AUDI"})
        honda = wb["HONDA"]
        rows = list(honda.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 2)
        headers = [c.value for c in honda[1]]
        self.assertIn("Price", headers)
        self.assertIn("VIN",   headers)
        self.assertIn("Lot Number", headers)

    def test_processed_file_skipped(self):
        shutil.copy(
            CSV_FIXTURES / "copart_price_with_in_progress.csv",
            self.work_dir / "copart_price_2026_01_02.csv",
        )
        processed = {"copart_price_2026_01_02.csv"}
        pending = build_workbook.find_pending_files(
            self.work_dir, date(2026, 6, 1), processed,
        )
        self.assertEqual(pending, [])

    def test_today_file_skipped(self):
        # Files dated today must NOT be imported (day hasn't ended)
        today = date.today().strftime("%Y_%m_%d")
        path  = self.work_dir / f"copart_price_{today}.csv"
        shutil.copy(CSV_FIXTURES / "copart_price_with_in_progress.csv", path)
        pending = build_workbook.find_pending_files(self.work_dir, date.today(), set())
        self.assertEqual(pending, [])

    def test_auction_date_normalized_on_import(self):
        """CSVs with raw IAAI dates get normalized on workbook import."""
        csv_path = self.work_dir / "iaai_price_2026_04_21.csv"
        csv_path.write_text(
            "Make,Model,Odometer,Price,Lot Number,Link,Auction Date,VIN\n"
            'HONDA,CR-V,15000,$1,111,"https://x","Tue Apr 21, 8:30am CDT",V1\n',
            encoding="utf-8",
        )
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        build_workbook.process_csv(csv_path, wb)

        ws = wb["HONDA"]
        headers = [c.value for c in ws[1]]
        date_i  = headers.index("Auction Date")
        self.assertEqual(ws.cell(row=2, column=date_i + 1).value,
                         "2026-04-21 13:30 UTC")

    def test_self_healing_normalizes_stale_workbook_cells(self):
        """Historical IAAI-format rows already in the workbook get rewritten."""
        wb_path = self.work_dir / "wb.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("HONDA")
        ws.append(["Make", "Model", "Auction Date", "Lot Number"])
        ws.append(["HONDA", "CR-V", "Tue Apr 21, 8:30am CDT", "11111111"])  # stale
        ws.append(["HONDA", "CR-V", "2026-04-22 14:00 UTC",    "22222222"])  # canonical
        wb.save(wb_path)

        wb2    = openpyxl.load_workbook(wb_path)
        healed = build_workbook.normalize_existing_auction_dates(wb2)
        self.assertEqual(healed, 1)

        self.assertEqual(wb2["HONDA"].cell(row=2, column=3).value,
                         "2026-04-21 13:30 UTC")
        self.assertEqual(wb2["HONDA"].cell(row=3, column=3).value,
                         "2026-04-22 14:00 UTC")

    def test_self_healing_noop_on_clean_workbook(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("AUDI")
        ws.append(["Make", "Auction Date"])
        ws.append(["AUDI", "2026-04-21 13:30 UTC"])
        self.assertEqual(build_workbook.normalize_existing_auction_dates(wb), 0)


if __name__ == "__main__":
    unittest.main()
