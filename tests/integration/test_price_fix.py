"""Integration tests for scripts/price_fix.py with FakeBidfaxClient."""

import csv
import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from tests._helpers import CSV_FIXTURES, ROOT  # noqa: F401

import price_fix
from clients.bidfax import FakeBidfaxClient


def _build_test_workbook(path: Path) -> None:
    """Create a tiny workbook with a Lot Number column across two sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("HONDA")
    ws1.append(["Make", "Model", "Price", "Lot Number", "Link", "VIN"])
    ws1.append(["HONDA", "CR-V", "$18,500", "11111111",
                '=HYPERLINK("https://old.example/one")', "OLDVIN1"])
    ws1.append(["HONDA", "CR-V", "In Progress", "22222222",
                '=HYPERLINK("https://old.example/two")', ""])

    ws2 = wb.create_sheet("AUDI")
    ws2.append(["Make", "Model", "Price", "Lot Number", "Link", "VIN"])
    ws2.append(["AUDI", "Q5", "In Progress", "33333333",
                '=HYPERLINK("https://old.example/three")', ""])

    wb.save(path)


def _build_test_html(path: Path) -> None:
    """Create a minimal HTML report with a filterable-table containing our lots."""
    html = """<!DOCTYPE html>
<html><body>
<table class="filterable-table main-table">
<thead><tr>
<th>Make<span class="sort-icon"></span></th>
<th>Model<span class="sort-icon"></span></th>
<th>Price<span class="sort-icon"></span></th>
<th>Lot Number<span class="sort-icon"></span></th>
<th>Link<span class="sort-icon"></span></th>
<th>VIN<span class="sort-icon"></span></th>
</tr></thead>
<tbody>
<tr data-model="CR-V">
<td>HONDA</td><td class="cell-model">CR-V</td>
<td class="cell-price">$18,500</td><td>11111111</td>
<td class="cell-link"><a href="https://old.example/one" target="_blank">View</a></td>
<td class="cell-vin">OLDVIN1</td>
</tr>
<tr data-model="CR-V">
<td>HONDA</td><td class="cell-model">CR-V</td>
<td class="cell-price">In Progress</td><td>22222222</td>
<td class="cell-link"><a href="https://old.example/two" target="_blank">View</a></td>
<td class="cell-vin"></td>
</tr>
<tr data-model="Q5">
<td>AUDI</td><td class="cell-model">Q5</td>
<td class="cell-price">In Progress</td><td>33333333</td>
<td class="cell-link"><a href="https://old.example/three" target="_blank">View</a></td>
<td class="cell-vin"></td>
</tr>
</tbody>
</table>
</body></html>"""
    path.write_text(html, encoding="utf-8")


class TestPriceFix(unittest.TestCase):
    def setUp(self):
        self._tmp     = tempfile.TemporaryDirectory()
        self.work_dir = Path(self._tmp.name)
        shutil.copy(
            CSV_FIXTURES / "copart_price_with_in_progress.csv",
            self.work_dir / "copart_price_2026_01_02.csv",
        )
        self.workbook_path = self.work_dir / "auction_results.xlsx"
        _build_test_workbook(self.workbook_path)
        self.html_path = self.work_dir / "index.html"
        _build_test_html(self.html_path)

    def tearDown(self):
        self._tmp.cleanup()

    def test_lookup_lots_filters_not_found(self):
        """Lots without a bidfax URL should be dropped from results."""
        fake = FakeBidfaxClient(responses={
            "22222222": ("$17,500", "VINB", "https://bidfax.info/honda/cr-v/b-vin-vinb.html"),
            # 33333333 has no response → no url → skipped
        })
        results = price_fix.lookup_lots(["22222222", "33333333"], delay=0,
                                        browser_port=None, client=fake)
        self.assertIn("22222222", results)
        self.assertNotIn("33333333", results)

    def test_fix_csv_workbook_html(self):
        results = {
            "22222222": ("$17,500", "VINB", "https://bidfax.info/honda/cr-v/b-vin-vinb.html"),
            "33333333": ("$14,000", "VINC", "https://bidfax.info/audi/q5/c-vin-vinc.html"),
        }

        # --- CSV ---
        total = price_fix.fix_csvs(self.work_dir, results)
        self.assertEqual(total, 2)
        with (self.work_dir / "copart_price_2026_01_02.csv").open() as fh:
            rows = {r["Lot Number"]: r for r in csv.DictReader(fh)}
        self.assertEqual(rows["22222222"]["Price"], "$17,500")
        self.assertEqual(rows["22222222"]["VIN"],   "VINB")
        self.assertIn("bidfax.info", rows["22222222"]["Link"])
        self.assertEqual(rows["33333333"]["Price"], "$14,000")

        # --- Workbook ---
        total = price_fix.fix_workbook(self.workbook_path, results)
        self.assertEqual(total, 2)
        wb = openpyxl.load_workbook(self.workbook_path)
        honda = wb["HONDA"]
        headers = [c.value for c in honda[1]]
        price_i = headers.index("Price") + 1
        vin_i   = headers.index("VIN") + 1
        link_i  = headers.index("Link") + 1
        # row 3 is the 22222222 row
        self.assertEqual(honda.cell(row=3, column=price_i).value, "$17,500")
        self.assertEqual(honda.cell(row=3, column=vin_i).value,   "VINB")
        self.assertIn("bidfax.info", honda.cell(row=3, column=link_i).value)

        # --- HTML ---
        total = price_fix.fix_html(self.html_path, results)
        self.assertEqual(total, 2)
        html = self.html_path.read_text()
        self.assertIn("$17,500", html)
        self.assertIn("VINB",    html)
        self.assertIn("$14,000", html)
        self.assertIn("VINC",    html)
        # Link cell now points to bidfax (cell-bidfax class)
        self.assertIn("cell-bidfax", html)
        self.assertIn("https://bidfax.info/honda/cr-v/b-vin-vinb.html", html)

    def test_fix_skips_unknown_lots(self):
        results = {"99999999": ("$0", "V", "https://bidfax.info/x/y/z.html")}
        self.assertEqual(price_fix.fix_csvs(self.work_dir, results), 0)
        self.assertEqual(price_fix.fix_workbook(self.workbook_path, results), 0)
        self.assertEqual(price_fix.fix_html(self.html_path, results), 0)


if __name__ == "__main__":
    unittest.main()
