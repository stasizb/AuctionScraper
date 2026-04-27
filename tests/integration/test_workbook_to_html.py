"""Integration tests for scripts/workbook_to_html.py using --no-bidfax and FakeBidfaxClient."""

import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from tests._helpers import CSV_FIXTURES, ROOT  # noqa: F401

import workbook_to_html
from clients.bidfax import FakeBidfaxClient


def _build_test_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("HONDA")
    ws.append(["Make", "Model", "Year", "Odometer", "Price", "Fuel Type",
               "Lot Number", "Link", "Auction Date", "Location",
               "Primary Damage", "VIN"])
    ws.append(["HONDA", "CR-V", 2024, "15,000", "$18,500", "Hybrid",
               "11111111", '=HYPERLINK("https://bidfax.info/honda/cr-v/one.html")',
               "2026-01-02 14:00 UTC", "CO - DENVER", "REAR END", "VIN111"])
    ws.append(["HONDA", "CR-V", 2023, "20,000", "$17,000", "Hybrid",
               "22222222", '=HYPERLINK("https://www.copart.com/lot/22222222/honda-cr-v")',
               "2026-01-02 15:00 UTC", "TX - DALLAS", "FRONT END", "VIN222"])
    wb.save(path)


class TestWorkbookToHtml(unittest.TestCase):
    def setUp(self):
        self._tmp     = tempfile.TemporaryDirectory()
        self.work_dir = Path(self._tmp.name)
        self.workbook_path = self.work_dir / "auction_results.xlsx"
        _build_test_workbook(self.workbook_path)

    def tearDown(self):
        self._tmp.cleanup()

    def _generate(self, client=None, search_dir: Path | None = None, today_date="2026_01_02"):
        out_dir = self.work_dir / "html_report"
        if out_dir.exists():
            shutil.rmtree(out_dir)
        out_dir.mkdir(parents=True)

        wb      = openpyxl.load_workbook(self.workbook_path)
        vins    = workbook_to_html._collect_vins(wb)
        vin_to_url = (workbook_to_html._lookup_bidfax_urls(
            vins, self.work_dir / "cache.json", delay=0, client=client,
        ) if client is not None else None)

        today_lots = workbook_to_html._load_today_lots(
            search_dir or self.work_dir, today_date,
        )
        html = workbook_to_html._build_html(wb, "Test", vin_to_url, today_lots)
        (out_dir / "index.html").write_text(html, encoding="utf-8")
        return (out_dir / "index.html").read_text()

    def test_render_with_bidfax_link_kept(self):
        """A row whose Link is already bidfax should render a 'Bidfax' button."""
        html = self._generate(client=FakeBidfaxClient())
        self.assertIn("cell-bidfax",                 html)
        self.assertIn("bidfax.info/honda/cr-v/one",  html)
        self.assertIn(">Bidfax<",                    html)

    def test_render_with_fallback_link(self):
        """A row whose Link is NOT bidfax renders as 'View' button."""
        html = self._generate(client=FakeBidfaxClient())
        self.assertIn("cell-link", html)
        self.assertIn(">View<",    html)

    def test_vin_lookup_resolves_to_bidfax_button(self):
        """Fake bidfax supplies a URL for VIN222 — Link cell should use it."""
        fake = FakeBidfaxClient(responses={
            "VIN222": ("$17,000", "VIN222", "https://bidfax.info/honda/cr-v/resolved-vin-vin222.html"),
        })
        html = self._generate(client=fake)
        self.assertIn("resolved-vin-vin222", html)

    def test_today_section_rendered_from_csv(self):
        """Today's search CSV should add a 'Today's Auctions' section."""
        shutil.copy(CSV_FIXTURES / "copart_search_2026_01_02.csv", self.work_dir)
        html = self._generate(client=FakeBidfaxClient(),
                              search_dir=self.work_dir, today_date="2026_01_02")
        self.assertIn("Today", html)
        # today's lot numbers from the fixture
        self.assertIn("11111111", html)

    def test_make_only_in_today_lots_gets_its_own_tab(self):
        """A Make freshly added to filters/ shows up in today's CSV but not
        yet in the workbook (no priced lots). It must still render a tab/panel
        — otherwise the user can't see today's new-make auctions until the
        next pricing pass runs."""
        # The fixture CSV has rows for HONDA + AUDI; only HONDA exists in the
        # test workbook. AUDI is the "new make" the workbook hasn't seen yet.
        shutil.copy(CSV_FIXTURES / "copart_search_2026_01_02.csv", self.work_dir)
        html = self._generate(client=FakeBidfaxClient(),
                              search_dir=self.work_dir, today_date="2026_01_02")
        # Tab present for the new make
        self.assertIn('data-target="AUDI"', html)
        self.assertIn(">AUDI<", html)
        # Panel exists with today's AUDI lot rendered as a today-only block
        self.assertIn('id="AUDI"', html)
        self.assertIn("33333333", html)  # AUDI Q5 lot from the fixture
        # Subtitle reflects the new make in the count (workbook had 1 make)
        self.assertIn("2 make(s)", html)

        # The new-make panel must look like the green "Today's Auctions"
        # section, not the blue main grid (regression: old code emitted only
        # main-table class so the rendered table inherited the blue header).
        audi_start = html.find('id="AUDI"')
        next_panel = html.find('class="tab-panel', audi_start + 1)
        audi_panel = html[audi_start: next_panel if next_panel != -1 else len(html)]
        self.assertIn("today-section",        audi_panel)
        self.assertIn("Today's Auctions",     audi_panel)
        self.assertIn("today-table",          audi_panel)

    def test_summary_section_rendered(self):
        html = self._generate(client=FakeBidfaxClient())
        self.assertIn("Summary by Model", html)
        self.assertIn("CR-V", html)

    def test_mobile_structure_elements_present(self):
        """Mobile features: data-field attrs, <details>, hamburger, mobile CSS."""
        html = self._generate(client=FakeBidfaxClient())

        # data-field / data-label on every <td> that mobile CSS targets
        self.assertIn('data-field="model"',          html)
        self.assertIn('data-field="year"',           html)
        self.assertIn('data-field="odometer"',       html)
        self.assertIn('data-field="lot-number"',     html)
        self.assertIn('data-field="price"',          html)
        self.assertIn('data-field="link"',           html)
        self.assertIn('data-field="primary-damage"', html)
        self.assertIn('data-field="vin"',            html)

        # Model filter and summary wrapped in <details> (open by default)
        self.assertIn('<details class="summary-section" open>', html)
        self.assertIn('<details class="model-filter" open>',    html)

        # Hamburger button present
        self.assertIn('id="mobile-menu-btn"', html)

    def test_mobile_css_and_js_constants(self):
        """Mobile rules and hooks present in the inline CSS / JS constants."""
        css = workbook_to_html.CSS
        js  = workbook_to_html.JS

        self.assertIn("@media (max-width: 768px)", css)
        self.assertIn('data-field="model"',        css)
        self.assertIn("mobile-menu-btn",           css)

        # 2-row card grid: Model/Price/Link on row 1;
        # Damage + Odometer on row 2 (Odo under Price, col 3 empty)
        self.assertIn('grid-template-areas', css)
        self.assertIn('"model price link"', css)
        self.assertIn('"dmg   odo   ."',     css)

        # Fields hidden on mobile cards include Year, VIN, Lot Number
        self.assertIn('data-field="year"', css)
        self.assertIn('data-field="vin"',  css)
        self.assertIn('data-field="lot-number"', css)

        # Price column: right-aligned + fixed width so it lines up across cards
        self.assertIn('grid-template-columns: 1fr 85px 70px', css)

        # Link: right-aligned via justify-self:end
        self.assertIn('justify-self: end', css)

        # Today's cards hide Price
        self.assertIn(".today-table tbody td[data-field=\"price\"] { display: none; }", css)

        # Hamburger menu: tabs hidden by default, drawer shown as list-style
        self.assertIn(".tab-strip { display: none; }", css)
        self.assertIn("body.menu-open .tab-strip",     css)
        self.assertIn("flex-direction: column",        css)  # drawer is a list, not a row

        self.assertIn("sortByAuctionDateDesc", js)
        self.assertIn("Auction Date",          js)
        self.assertIn("matchMedia",            js)
        self.assertIn("menu-open",             js)


if __name__ == "__main__":
    unittest.main()
