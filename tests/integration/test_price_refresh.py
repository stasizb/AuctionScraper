"""Integration tests for scripts/price_refresh.py with FakeBidfaxClient."""

import csv
import json
import shutil
import tempfile
import unittest
from pathlib import Path

import openpyxl

from tests._helpers import CSV_FIXTURES, ROOT  # noqa: F401

import price_refresh
from clients.bidfax import FakeBidfaxClient


class TestPriceRefresh(unittest.TestCase):
    def setUp(self):
        self._tmp     = tempfile.TemporaryDirectory()
        self.work_dir = Path(self._tmp.name)
        # Fixture has one row with a real price + two with In Progress
        shutil.copy(
            CSV_FIXTURES / "copart_price_with_in_progress.csv",
            self.work_dir / "copart_price_2026_01_02.csv",
        )

    def tearDown(self):
        self._tmp.cleanup()

    def test_refresh_updates_in_progress_rows(self):
        fake = FakeBidfaxClient(responses={
            "22222222": ("$17,500", "VINB", "https://bidfax.info/honda/cr-v/b-vin-vinb.html"),
            "33333333": ("$14,000", "VINC", "https://bidfax.info/audi/q5/c-vin-vinc.html"),
        })
        files = price_refresh._find_price_files(self.work_dir, auction="all")
        self.assertEqual(len(files), 1)

        file_data, pending = price_refresh._collect_pending(files)
        self.assertEqual(set(pending.keys()), {"22222222", "33333333"})

        cache_path = self.work_dir / "cache.json"
        results = price_refresh._fetch_prices(pending, cache_path, delay=0, client=fake)

        self.assertIn("22222222", results)
        self.assertIn("33333333", results)

        total, touched = price_refresh._apply_results(file_data, results)
        self.assertEqual(total, 2)
        self.assertEqual(touched, 1)

        with (self.work_dir / "copart_price_2026_01_02.csv").open() as fh:
            rows = list(csv.DictReader(fh))
        by_lot = {r["Lot Number"]: r for r in rows}
        self.assertEqual(by_lot["11111111"]["Price"], "$18,500")  # untouched
        self.assertEqual(by_lot["22222222"]["Price"], "$17,500")
        self.assertEqual(by_lot["22222222"]["VIN"],   "VINB")
        self.assertIn("bidfax.info", by_lot["22222222"]["Link"])
        self.assertEqual(by_lot["33333333"]["Price"], "$14,000")

        # Cache now contains the confirmed prices
        cache = json.loads(cache_path.read_text())
        self.assertIn("22222222", cache)
        self.assertIn("33333333", cache)

    def test_workbook_single_open_flow(self):
        """fix #5 — workbook should be opened once and saved once per run."""
        # Build a small workbook that has one In Progress row missing from the CSVs
        wb_path = self.work_dir / "auction_results.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("AUDI")
        ws.append(["Make", "Model", "Price", "Lot Number", "Link", "VIN"])
        # 33333333 is In Progress in BOTH csv AND workbook
        ws.append(["AUDI", "Q5", "In Progress", "33333333",
                   '=HYPERLINK("https://old/three")', ""])
        # 44444444 is In Progress only in the workbook (simulates CSV-already-refreshed)
        ws.append(["AUDI", "Q5", "In Progress", "44444444",
                   '=HYPERLINK("https://old/four")', ""])
        wb.save(wb_path)

        fake = FakeBidfaxClient(responses={
            "33333333": ("$14,000", "VINC", "https://bidfax.info/audi/q5/c-vin-vinc.html"),
            "44444444": ("$12,000", "VIND", "https://bidfax.info/audi/q5/d-vin-vind.html"),
        })

        # Simulate the main() flow, using the new single-open helpers
        files = price_refresh._find_price_files(self.work_dir, auction="all")
        file_data, pending = price_refresh._collect_pending(files)
        opened_wb, wb_pending = price_refresh._open_workbook_collect_pending(wb_path)
        self.assertIsNotNone(opened_wb)
        self.assertIn("44444444", wb_pending)  # workbook-only lot captured
        for lot, make in wb_pending.items():
            if lot not in pending:
                pending[lot] = (None, make)

        results = price_refresh._fetch_prices(
            pending, self.work_dir / "cache.json", delay=0, client=fake,
        )
        price_refresh._apply_results(file_data, results)
        updated = price_refresh._apply_to_open_workbook(opened_wb, results)
        opened_wb.save(wb_path)
        self.assertEqual(updated, 2)

        # Reload to verify on-disk state
        wb2 = openpyxl.load_workbook(wb_path)
        ws2 = wb2["AUDI"]
        rows = list(ws2.iter_rows(min_row=2, values_only=True))
        by_lot = {r[3]: r for r in rows}
        self.assertEqual(by_lot["33333333"][2], "$14,000")
        self.assertEqual(by_lot["44444444"][2], "$12,000")
        self.assertIn("bidfax.info", str(by_lot["33333333"][4]))

    def test_refresh_leaves_untouched_when_still_in_progress(self):
        # No responses → all lookups still In Progress → nothing changes
        fake = FakeBidfaxClient()
        files = price_refresh._find_price_files(self.work_dir, auction="all")
        file_data, pending = price_refresh._collect_pending(files)

        results = price_refresh._fetch_prices(pending, self.work_dir / "cache.json",
                                              delay=0, client=fake)
        total, _ = price_refresh._apply_results(file_data, results)
        self.assertEqual(total, 0)

        with (self.work_dir / "copart_price_2026_01_02.csv").open() as fh:
            rows = list(csv.DictReader(fh))
        by_lot = {r["Lot Number"]: r for r in rows}
        self.assertEqual(by_lot["22222222"]["Price"], "In Progress")
        self.assertEqual(by_lot["33333333"]["Price"], "In Progress")


if __name__ == "__main__":
    unittest.main()
