#!/usr/bin/env python3
"""
Consolidated bidfax pipeline — one browser session for all bidfax work.

Replaces the separate runs of:
  - bidfax_info.py copart  (Sale-Ended check + bidfax for newly-ended Copart lots)
  - bidfax_info.py iaai    (bidfax for today's IAAI lots)
  - price_refresh.py       (re-fetch In Progress prices in older price CSVs)

Phases (one asyncio.run, one fresh Chrome):
  1. Build a deduped bidfax queue with destination metadata per lot.
  2. Run bidfax lookup once (multi-tab, cache-aware).
  3. Distribute results: write/update CSV files and (optional) workbook.

The Chrome session is intentionally launched fresh (no --browser-port) so it
doesn't inherit cookies / Cloudflare score from the daily-run shared Chrome
— see TODO.md for the history that motivated this.

Usage:
    python scripts/bidfax_run.py --dir output \\
        --cache bidfax_cache.json --log logs/bidfax_deletions.json \\
        --workbook output/auction_results.xlsx
"""

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from clients import bidfax
from core.columns  import (AUCTION_DATE_COL, IN_PROGRESS, LOT_COL, MAKE_COL,
                           PRICE_COL, VIN_COL)
from core.csv_io   import (find_price_files, find_recent_search,
                           load_csv_dict, save_csv_dict)
from core.workbook import apply_result_to_row, resolve_columns

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required. Install with: pip install openpyxl")


# ---------------------------------------------------------------------------
# Job structure
# ---------------------------------------------------------------------------

@dataclass
class _NewRowDest:
    """Append a freshly-built row to <out_path> at write time."""
    out_path: Path
    src_row:  dict
    out_fieldnames: list[str]


@dataclass
class _UpdateRowDest:
    """Mutate an existing row dict in place (file gets re-saved at end)."""
    path: Path
    row:  dict


@dataclass
class _WorkbookRowDest:
    """A workbook row needing in-place update (resolved later)."""
    sheet: str
    lot:   str  # used to find the row when applying


@dataclass
class BidfaxJob:
    lot:          str
    make:         str
    destinations: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Staleness helpers
# ---------------------------------------------------------------------------

DEFAULT_STALE_DAYS  = 7

# IAAI auction dates land here as "YYYY-MM-DD HH:MM UTC". The trailing
# timezone abbrev is informational; we only need the date for staleness
# math, so strip whatever 3-4 letter zone tail is on it.
_TZ_SUFFIX_RE = re.compile(r"\s+[A-Z]{2,4}\s*$")


def _parse_auction_date(value) -> date | None:
    """Return the date portion of an Auction Date cell, or None if it
    can't be parsed. Defensive — a missing or malformed value must NOT
    cause a stale match (we'd rather keep a row than wrongly delete it)."""
    if value is None:
        return None
    s = _TZ_SUFFIX_RE.sub("", str(value).strip())
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _is_stale(auction_date_value, cutoff_days: int, today: date) -> bool:
    """True iff the auction date is more than `cutoff_days` days before
    `today`. Unparseable dates are never stale."""
    parsed = _parse_auction_date(auction_date_value)
    if parsed is None:
        return False
    return (today - parsed).days > cutoff_days


# ---------------------------------------------------------------------------
# Phase 1 — queue builders
# ---------------------------------------------------------------------------

def _build_output_fieldnames(src_fieldnames: list[str]) -> list[str]:
    cols = list(src_fieldnames)
    insert_at = cols.index("Odometer") + 1 if "Odometer" in cols else len(cols)
    cols.insert(insert_at, PRICE_COL)
    cols.append(VIN_COL)
    return cols


def _read_search_csv(path: Path) -> tuple[list[str], list[dict]]:
    fieldnames, rows = load_csv_dict(path)
    if not fieldnames:
        sys.exit(f"Input CSV is empty: {path}")
    if LOT_COL not in fieldnames:
        sys.exit(f"'{LOT_COL}' not found in {path.name}")
    return fieldnames, rows


def _filter_active_rows(rows: list[dict]) -> list[dict]:
    return [
        r for r in rows
        if str(r.get(MAKE_COL, "")).strip()
        and not str(r.get(MAKE_COL, "")).strip().startswith("#")
        and str(r.get(LOT_COL, "")).strip()
    ]


def _ensure_job(jobs: dict[str, BidfaxJob], lot: str, make: str) -> BidfaxJob:
    job = jobs.get(lot)
    if job is None:
        job = BidfaxJob(lot=lot, make=make)
        jobs[lot] = job
    elif not job.make and make:
        job.make = make
    return job


def collect_copart(
    jobs:        dict[str, BidfaxJob],
    input_path:  Path,
    output_path: Path,
    log_path:    Path,
    sale_ended:  dict[str, bool],
) -> tuple[list[dict], list[dict]]:
    """Build queue entries for today's Copart lots.

    Splits today's input into ended (queued for bidfax) and not-ended
    (deleted from input + appended to deletion log). Returns
    (active_rows, deleted_rows) for downstream output writing.
    """
    src_fieldnames, rows = _read_search_csv(input_path)
    active = _filter_active_rows(rows)
    out_fieldnames = _build_output_fieldnames(src_fieldnames)

    deleted: list[dict] = []
    kept:    list[dict] = []
    for row in active:
        url = str(row.get("Link", "")).strip()
        ended = sale_ended.get(url, False)
        if not ended:
            deleted.append(row)
            continue
        kept.append(row)
        lot  = str(row.get(LOT_COL, "")).strip()
        make = str(row.get(MAKE_COL, "")).strip()
        job  = _ensure_job(jobs, lot, make)
        job.destinations.append(_NewRowDest(output_path, row, out_fieldnames))

    if deleted:
        _append_deletion_log(log_path, input_path.name, deleted)
        deleted_lots = {str(r.get(LOT_COL, "")).strip() for r in deleted}
        _remove_from_input(input_path, src_fieldnames, rows, deleted_lots)

    return kept, deleted


def collect_iaai(
    jobs:        dict[str, BidfaxJob],
    input_path:  Path,
    output_path: Path,
) -> list[dict]:
    """Queue every active row from today's IAAI search."""
    src_fieldnames, rows = _read_search_csv(input_path)
    active = _filter_active_rows(rows)
    out_fieldnames = _build_output_fieldnames(src_fieldnames)
    for row in active:
        lot  = str(row.get(LOT_COL, "")).strip()
        make = str(row.get(MAKE_COL, "")).strip()
        job  = _ensure_job(jobs, lot, make)
        job.destinations.append(_NewRowDest(output_path, row, out_fieldnames))
    return active


def _classify_inprogress_row(
    row: dict, cutoff_days: int, today: date,
) -> tuple[str, str, bool] | None:
    """Inspect one In-Progress price-CSV row. Returns None if the row
    isn't In Progress or has no lot. Otherwise returns (lot, make, stale)
    so the caller can decide whether to queue it for refresh or prune it.
    """
    if row.get(PRICE_COL, "").strip() != IN_PROGRESS:
        return None
    lot  = str(row.get(LOT_COL, "")).strip()
    make = str(row.get(MAKE_COL, "")).strip()
    if not lot:
        return None
    stale = _is_stale(row.get(AUCTION_DATE_COL, ""), cutoff_days, today)
    return lot, make, stale


def collect_inprogress_csvs(
    jobs:    dict[str, BidfaxJob],
    files:   list[Path],
    *,
    cutoff_days: int = DEFAULT_STALE_DAYS,
    today:       date | None = None,
) -> tuple[dict[Path, tuple[list[str], list[dict]]], dict[Path, list[dict]]]:
    """Walk price CSVs. Queue fresh In-Progress rows for refresh; flag
    rows whose auction is older than `cutoff_days` for deletion (the
    caller drops them in phase 3).

    Returns (file_data, pruned) where:
      file_data — {path: (fieldnames, rows)} — all loaded files
      pruned    — {path: [row, ...]} — rows the file should NO LONGER contain
    """
    today = today or date.today()
    file_data: dict[Path, tuple[list[str], list[dict]]] = {}
    pruned:    dict[Path, list[dict]] = {}
    for path in files:
        fieldnames, rows = load_csv_dict(path)
        if PRICE_COL not in fieldnames:
            print(f"  [skip] no '{PRICE_COL}' column: {path.name}")
            continue
        file_data[path] = (fieldnames, rows)
        for row in rows:
            cls = _classify_inprogress_row(row, cutoff_days, today)
            if cls is None:
                continue
            lot, make, stale = cls
            if stale:
                pruned.setdefault(path, []).append(row)
                print(f"[*] Prune stale lot {lot} from {path.name} "
                      f"(auction {row.get(AUCTION_DATE_COL, '?')!s})")
                continue
            job = _ensure_job(jobs, lot, make)
            job.destinations.append(_UpdateRowDest(path, row))
    return file_data, pruned


def _scan_workbook_sheet(
    ws, jobs: dict[str, BidfaxJob],
    cutoff_days: int, today: date,
) -> list[int]:
    """Walk one sheet's In-Progress rows: queue fresh ones for refresh,
    return the 1-based row indices of stale ones for the caller to delete.
    Sheets without the required columns are skipped (empty list)."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header or LOT_COL not in header or PRICE_COL not in header:
        return []
    headers   = list(header)
    lot_i     = headers.index(LOT_COL)
    price_i   = headers.index(PRICE_COL)
    auct_i    = headers.index(AUCTION_DATE_COL) if AUCTION_DATE_COL in headers else None

    stale: list[int] = []
    for row_cells in ws.iter_rows(min_row=2):
        if str(row_cells[price_i].value or "").strip() != IN_PROGRESS:
            continue
        lot = str(row_cells[lot_i].value or "").strip()
        if not lot:
            continue
        ad_value = row_cells[auct_i].value if auct_i is not None else ""
        if _is_stale(ad_value, cutoff_days, today):
            stale.append(row_cells[0].row)
            print(f"[*] Prune stale lot {lot} from workbook sheet "
                  f"'{ws.title}' (auction {ad_value!s})")
            continue
        job = _ensure_job(jobs, lot, ws.title)
        job.destinations.append(_WorkbookRowDest(ws.title, lot))
    return stale


def collect_inprogress_workbook(
    jobs:        dict[str, BidfaxJob],
    workbook_path: Path | None,
    *,
    cutoff_days: int = DEFAULT_STALE_DAYS,
    today:       date | None = None,
):
    """Open the workbook (if any). Queue fresh In-Progress rows for
    refresh; return the wb plus a {sheet_name: [stale_row_indices]} map
    so phase 3 can delete those rows."""
    if workbook_path is None or not workbook_path.exists():
        return None, {}
    today = today or date.today()
    wb = openpyxl.load_workbook(workbook_path)
    pruned: dict[str, list[int]] = {}
    for ws in wb.worksheets:
        stale = _scan_workbook_sheet(ws, jobs, cutoff_days, today)
        if stale:
            pruned[ws.title] = stale
    return wb, pruned


# ---------------------------------------------------------------------------
# Deletion log + input rewrite (Copart-only)
# ---------------------------------------------------------------------------

def _append_deletion_log(log_path: Path, input_file: str, deleted_rows: list[dict]) -> None:
    existing = []
    if log_path.exists():
        try:
            existing = json.loads(log_path.read_text(encoding="utf-8"))
        except ValueError:
            existing = []
    existing.append({
        "run_date":      datetime.now().isoformat(),
        "input_file":    input_file,
        "deleted_items": deleted_rows,
    })
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text(json.dumps(existing, indent=2), encoding="utf-8")
    print(f"[*] Logged {len(deleted_rows)} deletion(s) → {log_path.name}")


def _remove_from_input(
    input_path:    Path,
    src_fieldnames: list[str],
    rows:           list[dict],
    deleted_lots:   set[str],
) -> None:
    kept = [r for r in rows if str(r.get(LOT_COL, "")).strip() not in deleted_lots]
    save_csv_dict(input_path, src_fieldnames, kept)
    print(f"[*] Removed {len(rows) - len(kept)} lot(s) from {input_path.name}")


# ---------------------------------------------------------------------------
# Phase 2 — bidfax lookup (one Chrome, multi-tab)
# ---------------------------------------------------------------------------

def run_bidfax_lookup(
    jobs:           dict[str, BidfaxJob],
    cache_path:     Path,
    delay:          float,
    max_concurrent: int,
    client:         bidfax.BidfaxClient | None = None,
) -> dict[str, tuple[str, str, str]]:
    """Query bidfax for every lot in `jobs`. Returns {lot: (price, vin, url)}.

    Cached lots short-circuit. Only non-In-Progress results land in the cache.
    Browser is launched fresh (no --browser-port) — see header comment.
    """
    if not jobs:
        return {}
    queries = list(jobs.keys())
    makes   = {lot: job.make for lot, job in jobs.items()}
    return bidfax.run_batch(
        queries, delay, cache_path,
        makes=makes, browser_port=None, client=client,
        max_concurrent=max_concurrent,
    )


def run_sale_ended_check(
    lot_urls:    list[str],
    client:      bidfax.BidfaxClient | None = None,
) -> dict[str, bool]:
    """Hit each Copart lot URL; return {url: True if 'Sale ended' visible}."""
    if not lot_urls:
        return {}
    real_client = client or bidfax.BrowserBidfaxClient(browser_port=None)
    return real_client.check_sale_ended_many(lot_urls)


# ---------------------------------------------------------------------------
# Phase 3 — distribute results
# ---------------------------------------------------------------------------

def _apply_to_existing_row(row: dict, price: str, vin: str, url: str) -> None:
    row[PRICE_COL] = price
    if vin:
        row[VIN_COL] = vin
    if url and row.get("Link", "").strip() != url:
        row["Link"] = url


def _build_new_row(src_row: dict, price: str, vin: str, url: str) -> dict:
    row = dict(src_row)
    row[PRICE_COL] = price
    row[VIN_COL]   = vin
    if url:
        row["Link"] = url
    return row


def _apply_jobs_to_destinations(
    jobs:    dict[str, BidfaxJob],
    results: dict[str, tuple[str, str, str]],
) -> tuple[dict[Path, tuple[list[str], list[dict]]], int]:
    """First pass over the queue: build per-output-path new-row buckets and
    mutate UpdateRowDest rows in place. Returns (new_rows_by_path, csv_updates)."""
    new_rows_by_path: dict[Path, tuple[list[str], list[dict]]] = {}
    csv_updates = 0
    for lot, job in jobs.items():
        price, vin, url = results.get(lot, (IN_PROGRESS, "", ""))
        for dest in job.destinations:
            if isinstance(dest, _NewRowDest):
                bucket = new_rows_by_path.setdefault(
                    dest.out_path, (dest.out_fieldnames, []),
                )
                bucket[1].append(_build_new_row(dest.src_row, price, vin, url))
            elif isinstance(dest, _UpdateRowDest):
                if price == IN_PROGRESS:
                    continue
                _apply_to_existing_row(dest.row, price, vin, url)
                csv_updates += 1
            # _WorkbookRowDest is handled in _apply_workbook below.
    return new_rows_by_path, csv_updates


def _write_new_outputs(
    new_rows_by_path: dict[Path, tuple[list[str], list[dict]]],
) -> None:
    for out_path, (fieldnames, rows) in new_rows_by_path.items():
        out_path.parent.mkdir(parents=True, exist_ok=True)
        save_csv_dict(out_path, fieldnames, rows)
        print(f"[+] Wrote {len(rows)} row(s) → {out_path}")


def _save_modified_existing_files(
    file_data:  dict[Path, tuple[list[str], list[dict]]],
    results:    dict[str, tuple[str, str, str]],
    csv_pruned: dict[Path, list[dict]],
) -> int:
    """Drop pruned rows from each file and re-save when anything changed
    (refresh update, prune, or both). Returns count of pruned rows."""
    pruned_total = 0
    for path, stale_rows in csv_pruned.items():
        if path not in file_data:
            continue
        fieldnames, rows = file_data[path]
        stale_ids = set(map(id, stale_rows))
        rows[:]   = [r for r in rows if id(r) not in stale_ids]
        pruned_total += len(stale_rows)

    for path, (fieldnames, rows) in file_data.items():
        had_prune  = path in csv_pruned
        had_update = any(_row_was_updated(r, results) for r in rows)
        if had_prune or had_update:
            save_csv_dict(path, fieldnames, rows)
            print(f"[+] Updated → {path.name}")
    return pruned_total


def distribute(
    jobs:        dict[str, BidfaxJob],
    results:     dict[str, tuple[str, str, str]],
    file_data:   dict[Path, tuple[list[str], list[dict]]],
    wb,
    *,
    csv_pruned:  dict[Path, list[dict]] | None = None,
    wb_pruned:   dict[str, list[int]]  | None = None,
) -> tuple[int, int, int]:
    """Write each lot's result into its destinations and apply pruning.

    Returns (csv_updates, wb_updates, pruned_rows_total)."""
    csv_pruned = csv_pruned or {}
    wb_pruned  = wb_pruned  or {}

    new_rows_by_path, csv_updates = _apply_jobs_to_destinations(jobs, results)
    _write_new_outputs(new_rows_by_path)
    pruned_csv = _save_modified_existing_files(file_data, results, csv_pruned)

    wb_updates = 0
    pruned_wb  = 0
    if wb is not None:
        wb_updates = _apply_workbook(wb, results)
        pruned_wb  = _delete_workbook_rows(wb, wb_pruned)

    return csv_updates, wb_updates, pruned_csv + pruned_wb


def _delete_workbook_rows(wb, wb_pruned: dict[str, list[int]]) -> int:
    """Remove the listed 1-based row indices from each sheet. Reverse-sort
    inside each sheet so deletions don't shift later indices."""
    total = 0
    for sheet_name, indices in wb_pruned.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for idx in sorted(indices, reverse=True):
            ws.delete_rows(idx)
            total += 1
    return total


def _row_was_updated(row: dict, results: dict[str, tuple]) -> bool:
    lot = str(row.get(LOT_COL, "")).strip()
    if lot not in results:
        return False
    price = results[lot][0]
    return price != IN_PROGRESS and row.get(PRICE_COL, "").strip() == price


def _apply_workbook(wb, results: dict[str, tuple[str, str, str]]) -> int:
    updated = 0
    for ws in wb.worksheets:
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
        cols = resolve_columns(headers)
        if cols is None:
            continue
        for row in ws.iter_rows(min_row=2):
            lot = str(row[cols.lot - 1].value or "").strip()
            if lot not in results:
                continue
            price, vin, url = results[lot]
            if price == IN_PROGRESS:
                continue
            apply_result_to_row(row, cols, price, vin, url)
            updated += 1
    return updated


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _resolve_input_paths(
    work_dir:    Path,
    today:       date,
    copart_date: str | None,
    iaai_date:   str | None,
) -> tuple[Path | None, Path | None, Path | None, Path | None]:
    """Pin Copart + IAAI input/output paths for this run. Each returns
    None when the corresponding search CSV can't be found."""
    yesterday = today - timedelta(days=1)
    cd  = copart_date or find_recent_search(work_dir, "copart", yesterday)
    id_ = iaai_date   or find_recent_search(work_dir, "iaai",   yesterday)
    copart_input  = work_dir / f"copart_search_{cd}.csv" if cd  else None
    copart_output = work_dir / f"copart_price_{cd}.csv"  if cd  else None
    iaai_input    = work_dir / f"iaai_search_{id_}.csv"  if id_ else None
    iaai_output   = work_dir / f"iaai_price_{id_}.csv"   if id_ else None
    return copart_input, copart_output, iaai_input, iaai_output


def _phase_copart(
    jobs:          dict[str, BidfaxJob],
    copart_input:  Path | None,
    copart_output: Path | None,
    log_path:      Path,
    bidfax_client: bidfax.BidfaxClient | None,
    copart_date:   str | None,
) -> None:
    if copart_input is None or not copart_input.exists():
        if copart_date is not None:
            print(f"[!] Copart input not found for date {copart_date}")
        return
    # _resolve_input_paths always returns input+output as a pair — when
    # input is set, output is set too. Guard for the type-checker.
    assert copart_output is not None
    _, raw_rows = _read_search_csv(copart_input)
    active_rows = _filter_active_rows(raw_rows)
    urls = [str(r.get("Link", "")).strip() for r in active_rows
            if str(r.get("Link", "")).strip()]
    print(f"[*] Sale-Ended check on {len(urls)} Copart lot(s)…")
    sale_ended = run_sale_ended_check(urls, bidfax_client)
    kept, deleted = collect_copart(jobs, copart_input, copart_output,
                                   log_path, sale_ended)
    print(f"[*] Copart: {len(kept)} ended → queued, "
          f"{len(deleted)} not-ended → removed")


def _phase_iaai(
    jobs:        dict[str, BidfaxJob],
    iaai_input:  Path | None,
    iaai_output: Path | None,
    iaai_date:   str | None,
) -> None:
    if iaai_input is None or not iaai_input.exists():
        if iaai_date is not None:
            print(f"[!] IAAI input not found for date {iaai_date}")
        return
    assert iaai_output is not None  # paired with iaai_input — see _resolve_input_paths
    active = collect_iaai(jobs, iaai_input, iaai_output)
    print(f"[*] IAAI: {len(active)} lot(s) → queued")


def _phase_refresh(
    jobs:              dict[str, BidfaxJob],
    work_dir:          Path,
    workbook_path:     Path | None,
    stale_cutoff_days: int,
    today:             date,
):
    """Scan price CSVs and workbook for In-Progress rows. Queue fresh
    ones, mark stale ones for deletion. Returns
    (file_data, csv_pruned, wb, wb_pruned)."""
    files = find_price_files(work_dir, "all")
    file_data, csv_pruned = collect_inprogress_csvs(
        jobs, files, cutoff_days=stale_cutoff_days, today=today,
    )
    refresh_count = sum(1 for j in jobs.values()
                        if any(isinstance(d, _UpdateRowDest) for d in j.destinations))
    pruned_csv_total = sum(len(v) for v in csv_pruned.values())
    print(f"[*] Refresh: {refresh_count} stale In Progress row(s) → queued, "
          f"{pruned_csv_total} stale row(s) → pruned (cutoff "
          f"{stale_cutoff_days} days)")

    wb, wb_pruned = collect_inprogress_workbook(
        jobs, workbook_path, cutoff_days=stale_cutoff_days, today=today,
    )
    pruned_wb_total = sum(len(v) for v in wb_pruned.values())
    if wb is not None:
        print(f"[*] Workbook: scanned for In Progress rows "
              f"({pruned_wb_total} stale row(s) → will prune)")
    return file_data, csv_pruned, wb, wb_pruned


def process(
    work_dir:        Path,
    cache_path:      Path,
    log_path:        Path,
    workbook_path:   Path | None,
    delay:           float,
    max_concurrent:  int,
    copart_date:     str | None = None,
    iaai_date:       str | None = None,
    bidfax_client:   bidfax.BidfaxClient | None = None,
    stale_cutoff_days: int = DEFAULT_STALE_DAYS,
    today:           date | None = None,
) -> None:
    today = today or date.today()
    copart_input, copart_output, iaai_input, iaai_output = _resolve_input_paths(
        work_dir, today, copart_date, iaai_date,
    )

    jobs: dict[str, BidfaxJob] = {}

    _phase_copart(jobs, copart_input, copart_output, log_path,
                  bidfax_client, copart_date)
    _phase_iaai(jobs, iaai_input, iaai_output, iaai_date)
    file_data, csv_pruned, wb, wb_pruned = _phase_refresh(
        jobs, work_dir, workbook_path, stale_cutoff_days, today,
    )

    if not jobs and not csv_pruned and not wb_pruned:
        print("[+] Nothing to query and nothing to prune — exiting.")
        return

    if jobs:
        print(f"\n[*] Bidfax queue: {len(jobs)} unique lot(s)")

    results = run_bidfax_lookup(
        jobs, cache_path, delay, max_concurrent, client=bidfax_client,
    ) if jobs else {}

    csv_updates, wb_updates, pruned_total = distribute(
        jobs, results, file_data, wb,
        csv_pruned=csv_pruned, wb_pruned=wb_pruned,
    )
    pruned_wb_total = sum(len(v) for v in wb_pruned.values())
    if wb is not None and (wb_updates or pruned_wb_total) and workbook_path is not None:
        wb.save(workbook_path)
        print(f"[+] Workbook saved ({wb_updates} updated, "
              f"{pruned_wb_total} pruned): {workbook_path.name}")

    print(f"\n[+] Done. CSV refresh updates: {csv_updates}, "
          f"workbook updates: {wb_updates}, pruned: {pruned_total}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Consolidated bidfax pipeline — replaces "
                    "bidfax_info + price_refresh.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--dir",         "-D", default="output",
                        help="Directory for input/output CSV files (default: output)")
    parser.add_argument("--cache",       "-c", default="bidfax_cache.json",
                        help="Cache file (default: bidfax_cache.json)")
    parser.add_argument("--log",         "-l", default="logs/bidfax_deletions.json",
                        help="Deletion log path (default: logs/bidfax_deletions.json)")
    parser.add_argument("--workbook",    "-w", default=None,
                        help="Excel workbook to update in-place (optional)")
    parser.add_argument("--copart-date", default=None,
                        help="Override Copart input date (yyyy_mm_dd)")
    parser.add_argument("--iaai-date",   default=None,
                        help="Override IAAI input date (yyyy_mm_dd)")
    parser.add_argument("--delay",       type=float, default=2.0,
                        help="Seconds between bidfax searches (default: 2.0)")
    parser.add_argument("--concurrent",  type=int,
                        default=bidfax.BIDFAX_TAB_CONCURRENCY,
                        help=f"Parallel bidfax tabs (default: "
                             f"{bidfax.BIDFAX_TAB_CONCURRENCY}; override via "
                             f"BIDFAX_TAB_CONCURRENCY / DEFAULT_TAB_CONCURRENCY "
                             f"env var or .env)")
    parser.add_argument("--stale-cutoff-days", type=int,
                        default=DEFAULT_STALE_DAYS,
                        help=f"In-Progress rows whose auction date is older "
                             f"than this many days are deleted from CSVs and "
                             f"workbook instead of re-queried "
                             f"(default: {DEFAULT_STALE_DAYS})")
    args = parser.parse_args()

    work_dir = Path(args.dir).resolve()
    process(
        work_dir          = work_dir,
        cache_path        = Path(args.cache),
        log_path          = Path(args.log),
        workbook_path     = Path(args.workbook) if args.workbook else None,
        delay             = args.delay,
        max_concurrent    = args.concurrent,
        copart_date       = args.copart_date,
        iaai_date         = args.iaai_date,
        stale_cutoff_days = args.stale_cutoff_days,
    )


if __name__ == "__main__":
    main()
