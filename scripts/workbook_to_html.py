#!/usr/bin/env python3
"""
Generate a beautiful HTML report from the auction results workbook.

Creates an output folder (default: html_report/) containing:
  index.html — one tab per Make, sortable + filterable tables
  style.css  — stylesheet
  script.js  — tab switching, column sorting, row filtering

The output folder is wiped and recreated on every run.

Usage:
    python workbook_to_html.py
    python workbook_to_html.py --workbook auction_results.xlsx --out html_report
    python workbook_to_html.py --title "My Auctions"
    python workbook_to_html.py --search-dir output --today-date 2026_04_10
"""

import argparse
import csv
import html as _html
import re
import shutil
import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found.  Install with:  pip install openpyxl")

from clients import bidfax
from core.dates import normalize_auction_date as _normalize_auction_date

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_HYPERLINK_RE    = re.compile(r'=HYPERLINK\("([^"]+)"', re.IGNORECASE)
_PRICE_RE        = re.compile(r'^\$([\d,]+)$')
_NUMERIC_COLS    = {"Year", "Odometer", "Price"}
_AUCTION_DATE_COL = "Auction Date"

_TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"


def _load_template(name: str) -> str:
    return (_TEMPLATE_DIR / name).read_text(encoding="utf-8")

CSS = _load_template("style.css")

JS  = _load_template("script.js")

# ---------------------------------------------------------------------------
# Bidfax lookup (browser-based)
# ---------------------------------------------------------------------------

_BIDFAX_DOMAIN = "bidfax.info"


def _row_link(row: tuple, link_idx: int) -> str:
    """Extract the resolved URL from a row's Link cell."""
    if link_idx < 0 or link_idx >= len(row):
        return ""
    raw = str(row[link_idx] or "").strip()
    m = _HYPERLINK_RE.match(raw)
    return m.group(1) if m else raw


def _vins_needing_lookup(ws) -> set[str]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    headers  = [str(h or "") for h in rows[0]]
    if "VIN" not in headers:
        return set()
    vin_idx  = headers.index("VIN")
    link_idx = headers.index("Link") if "Link" in headers else -1
    result: set[str] = set()
    for row in rows[1:]:
        vin = str(row[vin_idx] or "").strip() if vin_idx < len(row) else ""
        if not vin or vin.upper() == "NONE":
            continue
        if _BIDFAX_DOMAIN not in _row_link(row, link_idx):
            result.add(vin)
    return result


def _collect_vins(wb: openpyxl.Workbook) -> set[str]:
    vins: set[str] = set()
    for name in wb.sheetnames:
        vins |= _vins_needing_lookup(wb[name])
    return vins


def _lookup_bidfax_urls(
    vins: set[str],
    cache_path: Path,
    delay: float,
    browser_port: int | None = None,
    client: bidfax.BidfaxClient | None = None,
) -> dict[str, str]:
    return bidfax.run_batch_vins(
        sorted(vins), delay, cache_path,
        browser_port=browser_port, client=client,
    )


# ---------------------------------------------------------------------------
# Today's lots loader
# ---------------------------------------------------------------------------

def _load_today_lots(search_dir: Path, today_str: str) -> dict[str, list[dict]]:
    """Load copart + iaai search CSVs for today. Returns {MAKE_UPPER: [row_dicts]}."""
    result: dict[str, list[dict]] = {}
    for auction in ("copart", "iaai"):
        path = search_dir / f"{auction}_search_{today_str}.csv"
        if not path.exists():
            continue
        with path.open(newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                make = str(row.get("Make", "") or "").strip().upper()
                if make:
                    result.setdefault(make, []).append(dict(row))
    return result


# ---------------------------------------------------------------------------
# HTML cell / row helpers
# ---------------------------------------------------------------------------

def _extract_url(value) -> str | None:
    m = _HYPERLINK_RE.match(str(value or ""))
    return m.group(1) if m else None


_MODEL_MAX_LEN = 35


def _field_slug(header: str) -> str:
    """Normalise 'Lot Number' -> 'lot-number' for the data-field CSS hook."""
    return re.sub(r"[^a-z0-9]+", "-", header.lower()).strip("-")


def _td_attrs(header: str, extra: str = "") -> str:
    """Build the common data-field / data-label attrs for a <td>."""
    slug = _field_slug(header)
    return f'data-field="{slug}" data-label="{_html.escape(header)}" {extra}'.strip()


def _model_cell_html(raw: str) -> str:
    attrs = _td_attrs("Model", 'class="cell-model"')
    if len(raw) <= _MODEL_MAX_LEN:
        return f'<td {attrs}>{_html.escape(raw)}</td>'
    return (f'<td {attrs} title="{_html.escape(raw)}">'
            f'{_html.escape(raw[:_MODEL_MAX_LEN])}…</td>')


def _link_cell_html(raw: str) -> str:
    url = _extract_url(raw) or (raw if raw.startswith("http") else "")
    if not url:
        return f"<td {_td_attrs('Link')}></td>"
    klass = "cell-bidfax" if _BIDFAX_DOMAIN in url else "cell-link"
    label = "Bidfax"      if _BIDFAX_DOMAIN in url else "View"
    attrs = _td_attrs("Link", f'class="{klass}"')
    return f'<td {attrs}><a href="{_html.escape(url)}" target="_blank">{label}</a></td>'


def _cell_html(header: str, value) -> str:
    raw = "" if value is None else str(value).strip()

    if header == "Model":
        return _model_cell_html(raw)
    if header == "Link":
        return _link_cell_html(raw)
    if header == "Price":
        attrs = _td_attrs(header, 'class="cell-price"')
        return f'<td {attrs}>{_html.escape(raw)}</td>'
    if header == "VIN":
        attrs = _td_attrs(header, 'class="cell-vin"')
        return f'<td {attrs}>{_html.escape(raw)}</td>'
    if header in _NUMERIC_COLS:
        numeric = re.sub(r"[^\d.]", "", raw) or "0"
        attrs   = _td_attrs(header)
        return f'<td {attrs} data-raw="{_html.escape(numeric)}">{_html.escape(raw)}</td>'
    return f'<td {_td_attrs(header)}>{_html.escape(raw)}</td>'


# Fixed column widths — applied via <colgroup> so the main grid and the
# "Today's Auctions" grid share exactly the same layout, even when one of
# the tables has empty columns (e.g. Today's auctions have no Price yet).
# Values sum to ~1460px; with table-layout: fixed + width: 100% the
# browser scales them proportionally to fit the viewport.
_COL_WIDTHS = {
    "Make":           "100px",
    "Model":          "180px",
    "Year":            "55px",
    "Odometer":        "80px",
    "Price":           "85px",
    "Fuel Type":       "80px",
    "Lot Number":     "100px",
    "Link":            "75px",
    "Auction Date":   "160px",
    "Location":       "140px",
    "Primary Damage": "140px",
    "VIN":            "160px",
    "ACV":            "110px",
}


def _colgroup_html(headers: list) -> str:
    """Shared column widths for main + today tables so columns line up."""
    cols = "".join(
        f'<col style="width: {_COL_WIDTHS.get(h, "auto")}">' for h in headers
    )
    return f"<colgroup>{cols}</colgroup>"


def _thead_html(headers: list) -> str:
    cells = "".join(
        f'<th data-type="{"number" if h in _NUMERIC_COLS else "text"}">'
        f'{_html.escape(h)}<span class="sort-icon"></span></th>'
        for h in headers
    )
    return f"<thead><tr>{cells}</tr></thead>"


def _resolve_link(raw_value, vin: str, vin_to_url: dict | None) -> str:
    url = _extract_url(str(raw_value or "")) or (
        str(raw_value or "").strip() if str(raw_value or "").strip().startswith("http") else ""
    )
    if _BIDFAX_DOMAIN in url:
        return url
    if vin and vin_to_url:
        bidfax_url = vin_to_url.get(vin, "")
        if bidfax_url:
            return bidfax_url
    return url


# ---------------------------------------------------------------------------
# Summary section
# ---------------------------------------------------------------------------

def _summary_section_html(data_rows, headers: list[str]) -> str:
    """Static summary table: Model | Count | Avg Price (above the main grid)."""
    if "Model" not in headers or "Price" not in headers:
        return ""
    model_idx = headers.index("Model")
    price_idx = headers.index("Price")

    groups: dict[str, dict] = {}
    for row in data_rows:
        model = _model_key(str(row[model_idx].value or "").strip())
        if not model:
            continue
        price_raw = str(row[price_idx].value or "").strip()
        if model not in groups:
            groups[model] = {"count": 0, "prices": []}
        groups[model]["count"] += 1
        m = _PRICE_RE.match(price_raw)
        if m:
            groups[model]["prices"].append(float(m.group(1).replace(",", "")))

    if not groups:
        return ""

    rows_html = ""
    for model in sorted(groups):
        g   = groups[model]
        avg = f"${sum(g['prices']) / len(g['prices']):,.0f}" if g["prices"] else "—"
        rows_html += (
            f"<tr>"
            f"<td>{_html.escape(model)}</td>"
            f"<td>{g['count']}</td>"
            f'<td class="cell-price">{avg}</td>'
            f"</tr>"
        )

    # Open by default; JS closes it on mobile screens.
    return (
        '<details class="summary-section" open>'
        '<summary class="summary-label">Summary by Model</summary>'
        '<table class="summary-table no-sort">'
        "<thead><tr><th>Model</th><th>Count</th><th>Avg&nbsp;Price</th></tr></thead>"
        f"<tbody>{rows_html}</tbody>"
        "</table></details>"
    )


# ---------------------------------------------------------------------------
# Model filter widget
# ---------------------------------------------------------------------------

def _model_filter_html(models: list[str]) -> str:
    if not models:
        return ""
    chips = []
    chips.append(
        '<label class="model-chip all-chip">'
        '<input type="checkbox" class="model-cb" data-all="1" checked> All</label>'
    )
    for m in models:
        e = _html.escape(m)
        chips.append(
            f'<label class="model-chip">'
            f'<input type="checkbox" class="model-cb" value="{e}" checked> {e}</label>'
        )
    # Open by default; JS closes it on mobile screens (hamburger button re-opens).
    return (
        '<details class="model-filter" open>'
        '<summary class="model-filter-label">Model</summary>'
        f'<div class="model-chips">{"".join(chips)}</div>'
        '</details>'
    )


# ---------------------------------------------------------------------------
# Table body builders
# ---------------------------------------------------------------------------

def _get_cell_value(h: str, row, src_idx: dict, vin: str, vin_to_url: dict | None):
    i   = src_idx.get(h, -1)
    raw = row[i].value if i >= 0 else None
    if h == "Link":
        return _resolve_link(raw, vin, vin_to_url)
    if h == _AUCTION_DATE_COL and raw:
        return _normalize_auction_date(str(raw).strip())
    return raw


def _tbody_html(data_rows, src_headers: list, vin_idx, vin_to_url: dict | None) -> str:
    src_idx   = {h: i for i, h in enumerate(src_headers)}
    model_idx = src_idx.get("Model", -1)

    parts = []
    for row in data_rows:
        vin   = str(row[vin_idx].value or "").strip() if vin_idx is not None else ""
        model = _model_key(str(row[model_idx].value or "").strip() if model_idx >= 0 else "")
        cells = "".join(_cell_html(h, _get_cell_value(h, row, src_idx, vin, vin_to_url)) for h in src_headers)
        model_attr = f' data-model="{_html.escape(model)}"' if model else ""
        parts.append(f"<tr{model_attr}>{cells}</tr>")

    n = len(src_headers)
    parts.append(f'<tr class="no-results" style="display:none"><td colspan="{n}">No matching rows.</td></tr>')
    return f"<tbody>{''.join(parts)}</tbody>"


def _today_tbody_html(today_rows: list[dict], headers: list[str]) -> str:
    """Render today's search CSV rows using the workbook headers (Price/VIN left empty)."""
    parts = []
    for row in today_rows:
        model = _model_key(str(row.get("Model", "") or "").strip())
        cells = []
        for h in headers:
            if h in ("Price", "VIN"):
                val = ""
            else:
                val = str(row.get(h, "") or "").strip()
                if h == _AUCTION_DATE_COL:
                    val = _normalize_auction_date(val)
            cells.append(_cell_html(h, val))
        model_attr = f' data-model="{_html.escape(model)}"' if model else ""
        parts.append(f"<tr{model_attr}>{''.join(cells)}</tr>")

    n = len(headers)
    parts.append(f'<tr class="no-results" style="display:none"><td colspan="{n}">No matching rows.</td></tr>')
    return f"<tbody>{''.join(parts)}</tbody>"


# ---------------------------------------------------------------------------
# Model key helper
# ---------------------------------------------------------------------------

def _model_key(model: str) -> str:
    """Return the first word of a model string for grouping/filtering.

    e.g. 'GLE 350 4MATIC' → 'GLE', 'CR-V HYBRID' → 'CR-V', 'Q5' → 'Q5'
    """
    return model.split()[0] if model.strip() else ""


# ---------------------------------------------------------------------------
# Today-only panel builder (no workbook)
# ---------------------------------------------------------------------------

def _today_only_panel_content(today_rows: list[dict]) -> tuple[str, int]:
    """Build a panel containing only today's-auction rows.

    Used for makes that have no workbook sheet yet (freshly added to filters/),
    and as the sole content when the whole workbook is missing. The grid is
    styled as the green 'Today's Auctions' section because every row in it is
    a today-auction lot — there's no historical priced data to put under a
    blue main grid.
    """
    if not today_rows:
        return "<p>No data.</p>", 0

    headers = list(today_rows[0].keys())

    models = sorted({
        _model_key(str(r.get("Model", "") or "").strip())
        for r in today_rows
        if _model_key(str(r.get("Model", "") or "").strip())
    })

    filter_html = _model_filter_html(models)

    parts = []
    for row in today_rows:
        model = _model_key(str(row.get("Model", "") or "").strip())
        cells = []
        for h in headers:
            val = str(row.get(h, "") or "").strip()
            if h == _AUCTION_DATE_COL:
                val = _normalize_auction_date(val)
            cells.append(_cell_html(h, val))
        model_attr = f' data-model="{_html.escape(model)}"' if model else ""
        parts.append(f"<tr{model_attr}>{''.join(cells)}</tr>")

    n = len(headers)
    parts.append(f'<tr class="no-results" style="display:none"><td colspan="{n}">No matching rows.</td></tr>')
    tbody = f"<tbody>{''.join(parts)}</tbody>"

    table = (
        f'<table class="filterable-table main-table today-table">'
        f"{_colgroup_html(headers)}"
        f"{_thead_html(headers)}"
        f"{tbody}"
        f"</table>"
    )

    today_section = (
        '<div class="today-section">'
        '<div class="today-section-header">'
        f'<h3>Today\'s Auctions</h3>'
        f'<span class="today-badge">{len(today_rows)}</span>'
        "</div>"
        f'<div class="table-wrap">{table}</div>'
        "</div>"
    )

    content = filter_html + today_section
    return content, len(today_rows)


# ---------------------------------------------------------------------------
# Per-panel builder
# ---------------------------------------------------------------------------

def _extract_models(data_rows, headers: list[str], today_rows: list[dict]) -> list[str]:
    models: set[str] = set()
    if "Model" in headers:
        idx = headers.index("Model")
        for row in data_rows:
            val = _model_key(str(row[idx].value or "").strip())
            if val:
                models.add(val)
    for row in today_rows:
        val = _model_key(str(row.get("Model", "") or "").strip())
        if val:
            models.add(val)
    return sorted(models)


def _ws_to_panel_content(
    ws,
    vin_to_url: dict | None,
    today_rows: list[dict],
) -> tuple[str, int]:
    """Build the full HTML content for one tab panel. Returns (html, row_count)."""
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return "<p>No data.</p>", 0

    headers   = [str(c.value or "") for c in rows[0]]
    data_rows = rows[1:]
    vin_idx   = headers.index("VIN") if "VIN" in headers else None

    models = _extract_models(data_rows, headers, today_rows)

    filter_html  = _model_filter_html(models)
    summary_html = _summary_section_html(data_rows, headers)

    colgroup = _colgroup_html(headers)

    main_table = (
        f'<table class="filterable-table main-table">'
        f"{colgroup}"
        f"{_thead_html(headers)}"
        f"{_tbody_html(data_rows, headers, vin_idx, vin_to_url)}"
        f"</table>"
    )

    today_section = ""
    if today_rows:
        today_table = (
            f'<table class="filterable-table today-table">'
            f"{colgroup}"
            f"{_thead_html(headers)}"
            f"{_today_tbody_html(today_rows, headers)}"
            f"</table>"
        )
        today_section = (
            '<div class="today-section">'
            '<div class="today-section-header">'
            f'<h3>Today\'s Auctions</h3>'
            f'<span class="today-badge">{len(today_rows)}</span>'
            "</div>"
            f'<div class="table-wrap">{today_table}</div>'
            "</div>"
        )

    # Today's Auctions goes first — active lots the user is deciding on
    # right now are more useful than the historical grid underneath.
    content = (
        filter_html
        + summary_html
        + today_section
        + f'<div class="table-wrap">{main_table}</div>'
    )
    return content, len(data_rows)


# ---------------------------------------------------------------------------
# Full page builder
# ---------------------------------------------------------------------------

def _build_html(
    wb: openpyxl.Workbook | None,
    title: str,
    vin_to_url: dict | None,
    today_lots: dict[str, list[dict]],
) -> str:
    tab_btns = []
    panels   = []

    if wb is not None:
        for i, name in enumerate(wb.sheetnames):
            make_upper  = name.upper()
            today_rows  = today_lots.get(make_upper, [])
            panel_html, count = _ws_to_panel_content(wb[name], vin_to_url, today_rows)
            safe_id = re.sub(r"\W+", "_", name)
            active  = "active" if i == 0 else ""

            tab_btns.append(
                f'<button class="tab-btn {active}" data-target="{safe_id}">'
                f'{_html.escape(name)}<span class="badge">{count}</span></button>'
            )
            panels.append(
                f'<div class="tab-panel {active}" id="{safe_id}">'
                f"{panel_html}</div>"
            )

        # Makes that appear in today's auction CSV but not yet in the workbook
        # (e.g. a Make freshly added to filters/, no priced lots yet) would
        # otherwise have no tab. Render them as today-only panels so users
        # can see the new lots immediately, before any pricing pass runs.
        sheet_makes_upper = {n.upper() for n in wb.sheetnames}
        new_makes = sorted(m for m in today_lots if m not in sheet_makes_upper)
        for make in new_makes:
            today_rows = today_lots[make]
            panel_html, count = _today_only_panel_content(today_rows)
            safe_id = re.sub(r"\W+", "_", make)
            tab_btns.append(
                f'<button class="tab-btn" data-target="{safe_id}">'
                f'{_html.escape(make)}<span class="badge">{count}</span></button>'
            )
            panels.append(
                f'<div class="tab-panel" id="{safe_id}">{panel_html}</div>'
            )

        total      = sum(wb[n].max_row - 1 for n in wb.sheetnames)
        make_count = len(wb.sheetnames) + len(new_makes)
        subtitle   = f"{total} vehicle(s) &nbsp;·&nbsp; {make_count} make(s)"
    else:
        for i, make in enumerate(sorted(today_lots.keys())):
            today_rows = today_lots[make]
            panel_html, count = _today_only_panel_content(today_rows)
            safe_id = re.sub(r"\W+", "_", make)
            active  = "active" if i == 0 else ""

            tab_btns.append(
                f'<button class="tab-btn {active}" data-target="{safe_id}">'
                f'{_html.escape(make)}<span class="badge">{count}</span></button>'
            )
            panels.append(
                f'<div class="tab-panel {active}" id="{safe_id}">'
                f"{panel_html}</div>"
            )

        total_today = sum(len(v) for v in today_lots.values())
        subtitle    = f"{total_today} vehicle(s) &nbsp;·&nbsp; {len(today_lots)} make(s) — Today's lots only"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{_html.escape(title)}</title>
  <link rel="stylesheet" href="style.css">
</head>
<body>
<header>
  <h1>{_html.escape(title)}</h1>
  <span class="subtitle">{subtitle}</span>
</header>
<div class="container">
  <div class="tab-strip">{"".join(tab_btns)}</div>
  <div class="toolbar">
    <button id="mobile-menu-btn" class="mobile-menu-btn" aria-label="Toggle model filter">☰</button>
    <input id="search-input" type="search" placeholder="Filter visible table…">
    <span class="row-count" id="row-count"></span>
  </div>
  {"".join(panels)}
</div>
<script src="script.js"></script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    today_default = date.today().strftime("%Y_%m_%d")

    parser = argparse.ArgumentParser(
        description="Generate an HTML report from the auction results workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--workbook",     "-w", default="auction_results.xlsx",
                        help="Source workbook (default: auction_results.xlsx)")
    parser.add_argument("--out",          "-o", default="html_report",
                        help="Output folder (default: html_report)")
    parser.add_argument("--title",        "-t", default="Auction Results",
                        help="Page title (default: Auction Results)")
    parser.add_argument("--search-dir",   "-s", default=None,
                        help="Directory with today's search CSVs (default: workbook directory)")
    parser.add_argument("--today-date",         default=today_default,
                        help=f"Date of today's search files yyyy_mm_dd (default: {today_default})")
    parser.add_argument("--no-bidfax",          action="store_true",
                        help="Skip Bidfax VIN lookup")
    parser.add_argument("--bidfax-cache",       default="bidfax_cache.json",
                        help="Cache file for bidfax lookups (default: bidfax_cache.json)")
    parser.add_argument("--bidfax-delay",  type=float, default=2.0,
                        help="Seconds between bidfax requests (default: 2.0)")
    parser.add_argument("--browser-port", type=int, default=None,
                        help="Connect to a running Chrome on this port instead of launching one")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    search_dir    = Path(args.search_dir) if args.search_dir else workbook_path.parent

    out_dir = Path(args.out)
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True)
    print(f"[*] Output folder : {out_dir.resolve()}")

    if not workbook_path.exists():
        print(f"[!] Workbook not found: {workbook_path} — skipping workbook conversion")
        wb         = None
        vin_to_url = None
    else:
        print(f"[*] Loading       : {workbook_path}")
        wb         = openpyxl.load_workbook(workbook_path)
        vins       = _collect_vins(wb)
        cache_path = Path(args.bidfax_cache)
        vin_to_url = None if args.no_bidfax else _lookup_bidfax_urls(
            vins, cache_path, args.bidfax_delay, browser_port=args.browser_port
        )

    (out_dir / "style.css").write_text(CSS, encoding="utf-8")
    (out_dir / "script.js").write_text(JS,  encoding="utf-8")
    print("[+] style.css  written")
    print("[+] script.js  written")

    print(f"[*] Loading today's lots from: {search_dir} (date: {args.today_date})")
    today_lots  = _load_today_lots(search_dir, args.today_date)
    total_today = sum(len(v) for v in today_lots.values())
    print(f"[*] Today's lots  : {total_today} across {len(today_lots)} make(s)")

    html_content = _build_html(wb, args.title, vin_to_url, today_lots)
    (out_dir / "index.html").write_text(html_content, encoding="utf-8")
    sheets_info = f"{len(wb.sheetnames)} sheet(s)" if wb is not None else "today-only"
    print(f"[+] index.html written  ({sheets_info})")
    print(f"\n[+] Done → {(out_dir / 'index.html').resolve()}")


if __name__ == "__main__":
    main()
