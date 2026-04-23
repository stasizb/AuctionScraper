#!/usr/bin/env python3
"""
Refresh "In Progress" prices in auction price CSV files.

Scans the current directory for files matching:
    <auction>_price_<yyyy>_<mm>_<dd>.csv  (iaai or copart)

For each file, rows where Price is "In Progress" are re-queried on bidfax.info.
Rows with a confirmed price are updated in-place. Cache is consulted first;
only confirmed (non-"In Progress") results are stored in cache.

USAGE:
    python price_refresh.py
    python price_refresh.py --auction iaai
    python price_refresh.py --dir /path/to/csvs --cache bidfax_cache.json
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from clients import bidfax
from core.columns  import IN_PROGRESS, LOT_COL, MAKE_COL, PRICE_COL, VIN_COL
from core.csv_io   import PRICE_FILE_PATTERN, find_price_files, load_csv_dict, save_csv_dict
from core.workbook import apply_result_to_row, resolve_columns

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found. Install with:  pip install openpyxl")

# Backwards-compatible aliases for tests and any external imports
FILE_PATTERN      = PRICE_FILE_PATTERN
_find_price_files = find_price_files
_load_csv         = load_csv_dict
_save_csv         = save_csv_dict


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _collect_pending(
    files: list[Path],
) -> tuple[dict[Path, tuple[list[str], list[dict]]], dict[str, tuple[Path, str]]]:
    """Load all price CSVs and collect lots with In Progress status.

    Returns (file_data, pending) where:
      file_data — {path: (fieldnames, rows)}
      pending   — {lot: (path, make)}
    """
    file_data: dict[Path, tuple[list[str], list[dict]]] = {}
    pending:   dict[str, tuple[Path, str]]               = {}

    for path in files:
        fieldnames, rows = _load_csv(path)
        if PRICE_COL not in fieldnames:
            print(f"  [skip] No '{PRICE_COL}' column: {path.name}")
            continue
        file_data[path] = (fieldnames, rows)
        for row in rows:
            if row.get(PRICE_COL, "").strip() != IN_PROGRESS:
                continue
            lot  = str(row.get(LOT_COL, "")).strip()
            make = str(row.get(MAKE_COL, "")).strip()
            if lot:
                pending[lot] = (path, make)

    return file_data, pending


def _fetch_prices(
    pending: dict[str, tuple[Path, str]],
    cache_path: Path,
    delay: float,
    browser_port: int | None = None,
    client: bidfax.BidfaxClient | None = None,
    max_concurrent: int = 1,
) -> dict[str, tuple]:
    """Return confirmed (price, vin, url) for each lot, using cache where available."""
    cache    = bidfax.load_cache(cache_path)
    cached   = {lot: cache[lot] for lot in pending if lot in cache}
    to_fetch = [lot for lot in pending if lot not in cache]

    if cached:
        print(f"[*] {len(cached)} already in cache, {len(to_fetch)} need bidfax lookup")
    else:
        print(f"[*] {len(to_fetch)} need bidfax lookup")

    fetched: dict[str, tuple] = {}
    if to_fetch:
        makes   = {lot: pending[lot][1] for lot in to_fetch}
        results = bidfax.run_batch(
            to_fetch, delay, cache_path,
            makes=makes, browser_port=browser_port, client=client,
            max_concurrent=max_concurrent,
        )
        fetched = {lot: v for lot, v in results.items() if v[0] != IN_PROGRESS}

    return {**cached, **fetched}


def _update_row(row: dict, price: str, vin: str, url: str) -> None:
    """Apply a confirmed bidfax result to a single CSV row in-place."""
    row[PRICE_COL] = price
    if vin:
        row[VIN_COL] = vin
    if url and row.get("Link", "").strip() != url:
        row["Link"] = url


def _apply_to_file(
    path: Path,
    fieldnames: list[str],
    rows: list[dict],
    all_results: dict[str, tuple],
) -> int:
    """Update In Progress rows in one file. Returns count of rows changed."""
    updated = 0
    for row in rows:
        if row.get(PRICE_COL, "").strip() != IN_PROGRESS:
            continue
        lot = str(row.get(LOT_COL, "")).strip()
        if lot not in all_results:
            continue
        _update_row(row, *all_results[lot])
        updated += 1

    if updated:
        _save_csv(path, fieldnames, rows)
        print(f"  [+] Updated: {path.name}")
    return updated


def _sheet_in_progress(ws) -> dict[str, str]:
    """Return {lot: make} for every In Progress row in one worksheet."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}
    headers = list(header_row)
    if LOT_COL not in headers or PRICE_COL not in headers:
        return {}
    lot_i   = headers.index(LOT_COL)
    price_i = headers.index(PRICE_COL)
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[price_i] or "").strip() != IN_PROGRESS:
            continue
        lot = str(row[lot_i] or "").strip()
        if lot:
            result[lot] = ws.title  # sheet name = make
    return result


def _open_workbook_collect_pending(workbook_path: Path):
    """Open the workbook (writable) and collect {lot: make} of In Progress rows.

    Returns (wb, pending). `wb` is None if the workbook does not exist. The
    caller is responsible for saving and closing the workbook — this replaces
    the old pair of functions that opened the file twice (read-only then
    writable).
    """
    if not workbook_path.exists():
        return None, {}
    wb = openpyxl.load_workbook(workbook_path)
    pending: dict[str, str] = {}
    for ws in wb.worksheets:
        pending.update(_sheet_in_progress(ws))
    return wb, pending


def _apply_results(
    file_data: dict[Path, tuple[list[str], list[dict]]],
    all_results: dict[str, tuple],
) -> tuple[int, int]:
    """Write confirmed prices back into the CSV files.

    Returns (total_updated, updated_files).
    """
    total_updated = updated_files = 0
    for path, (fieldnames, rows) in file_data.items():
        n = _apply_to_file(path, fieldnames, rows, all_results)
        total_updated += n
        if n:
            updated_files += 1
    return total_updated, updated_files


def _apply_to_open_workbook(wb, all_results: dict[str, tuple]) -> int:
    """Apply confirmed results to an already-open workbook. Does not save."""
    total_updated = 0
    for ws in wb.worksheets:
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
        cols = resolve_columns(headers)
        if cols is None:
            continue
        for row in ws.iter_rows(min_row=2):
            lot = str(row[cols.lot - 1].value or "").strip()
            if lot not in all_results:
                continue
            price, vin, url = all_results[lot]
            apply_result_to_row(row, cols, price, vin, url)
            total_updated += 1
    return total_updated


def _update_workbook(workbook_path: Path, all_results: dict[str, tuple]) -> int:
    """Legacy API: open, update, save in one call. Kept for backward compatibility.

    Prefer `_open_workbook_collect_pending` + `_apply_to_open_workbook` when you
    need the pending-row scan — that avoids opening the file twice.
    """
    if not workbook_path.exists():
        print(f"  [skip] Workbook not found: {workbook_path}")
        return 0
    wb = openpyxl.load_workbook(workbook_path)
    total_updated = _apply_to_open_workbook(wb, all_results)
    if total_updated:
        wb.save(workbook_path)
        print(f"  [+] Workbook updated ({total_updated} row(s)): {workbook_path.name}")
    else:
        print("  [*] No workbook rows matched.")
    return total_updated


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Refresh In Progress prices in auction price CSV files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--auction",  "-a", default="all",
                        help="Auction name: copart, iaai, or all (default: all)")
    parser.add_argument("--dir",      "-d", default=".",
                        help="Directory to scan for CSV files (default: current dir)")
    parser.add_argument("--cache",    "-c", default="bidfax_cache.json",
                        help="Cache file for bidfax lookups (default: bidfax_cache.json)")
    parser.add_argument("--workbook", "-w", default=None,
                        help="Excel workbook to update in-place (optional)")
    parser.add_argument("--delay",        default=2.0, type=float,
                        help="Seconds between bidfax searches (default: 2.0)")
    parser.add_argument("--browser-port", type=int, default=None,
                        help="Connect to a running Chrome on this port instead of launching one")
    parser.add_argument("--concurrent", type=int, default=1,
                        help="Parallel bidfax tabs (default: 1 = sequential; experimental)")
    args = parser.parse_args()

    work_dir = Path(args.dir).resolve()
    files    = _find_price_files(work_dir, args.auction)

    if not files:
        print(f"[*] No matching price CSV files found in {work_dir}")
        return

    print(f"[*] Found {len(files)} file(s) to scan")

    file_data, pending = _collect_pending(files)

    # Open the workbook once (writable) and collect In Progress rows from it.
    # Covers the case where CSVs were already updated but the workbook wasn't.
    wb      = None
    wb_path = Path(args.workbook) if args.workbook else None
    if wb_path:
        wb, wb_pending = _open_workbook_collect_pending(wb_path)
        for lot, make in wb_pending.items():
            if lot not in pending:
                pending[lot] = (None, make)  # type: ignore[assignment]

    if not pending:
        print("[+] No In Progress rows found — nothing to refresh.")
        return

    print(f"[*] {len(pending)} lot(s) with In Progress status")

    all_results = _fetch_prices(
        pending, Path(args.cache), args.delay,
        browser_port=args.browser_port, max_concurrent=args.concurrent,
    )

    if not all_results:
        print("[*] No confirmed prices retrieved.")
        return

    if file_data:
        total_updated, updated_files = _apply_results(file_data, all_results)
        print(f"\n[+] Refreshed {total_updated} row(s) across {updated_files} file(s)")

    _save_workbook_results(wb, wb_path, all_results)


def _save_workbook_results(wb, wb_path: Path | None, all_results: dict[str, tuple]) -> None:
    """Apply confirmed results to an already-open workbook and save it."""
    if wb is None or wb_path is None:
        return
    print(f"\n[*] Propagating to workbook: {wb_path.name}")
    wb_updated = _apply_to_open_workbook(wb, all_results)
    if wb_updated:
        wb.save(wb_path)
        print(f"  [+] Workbook updated ({wb_updated} row(s)): {wb_path.name}")
    else:
        print("  [*] No workbook rows matched.")


if __name__ == "__main__":
    main()
