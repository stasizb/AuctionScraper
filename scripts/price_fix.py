#!/usr/bin/env python3
"""
price_fix.py — Re-fetch bidfax data for specific lot numbers and overwrite
every matching row in the output files.

Some lots end up with the wrong bidfax Link / Price / VIN because the
initial search matched a different vehicle. Give this script the affected
lot numbers and it will:

  1. Open one Chrome session on bidfax.info and search each lot.
  2. For every lot that is found, overwrite Link / Price / VIN wherever the
     lot appears in:
       - output/<auction>_price_<date>.csv files   (any number of matches)
       - output/auction_results.xlsx                (across every sheet)
       - output/html_report/index.html              (across every table)

Lots that are not found on bidfax are reported and skipped.

Usage:
    python price_fix.py --lots "44428368, 44428369, 44428360"
    python price_fix.py --lots 44428368 --dir output
    python price_fix.py --lots 44428368 --browser-port 9222
"""

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from clients import bidfax
from core.columns  import LINK_COL, LOT_COL, MAKE_COL, PRICE_COL, VIN_COL
from core.csv_io   import PRICE_FILE_PATTERN, load_csv_dict, save_csv_dict
from core.workbook import apply_result_to_row, resolve_columns

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found. Install with:  pip install openpyxl")

try:
    from bs4 import BeautifulSoup, NavigableString
    _BS4_OK = True
except ImportError:
    _BS4_OK = False


_BIDFAX_DOMAIN = "bidfax.info"


# ---------------------------------------------------------------------------
# CLI helpers
# ---------------------------------------------------------------------------

def _parse_lots(arg: str) -> list[str]:
    """Split '44428368, 44428369; 44428360' into ['44428368', '44428369', '44428360']."""
    return [p.strip() for p in re.split(r"[,;]", arg) if p.strip()]


# ---------------------------------------------------------------------------
# Bidfax lookup
# ---------------------------------------------------------------------------

def find_makes_for_lots(directory: Path, lots: list[str]) -> dict[str, str]:
    """Scan existing price CSVs in `directory` for each lot's Make.

    The make is needed for bidfax URL-make validation — without it, bidfax can
    return a totally different vehicle (e.g. Audi Q5 lot 41613606 silently
    pulled a Nissan Leaf result) and we'd happily overwrite the CSVs with it.
    Returns {lot: make_uppercase}; lots not found in any CSV are omitted.
    """
    wanted  = set(lots)
    makes: dict[str, str] = {}
    for path in sorted(directory.glob("*.csv")):
        if not PRICE_FILE_PATTERN.match(path.name) or not wanted:
            continue
        _fields, rows = load_csv_dict(path)
        for row in rows:
            lot = str(row.get(LOT_COL, "")).strip()
            if lot in wanted and not makes.get(lot):
                make = str(row.get(MAKE_COL, "")).strip()
                if make:
                    makes[lot] = make
        wanted -= set(makes)
    return makes


def lookup_lots(
    lots: list[str],
    delay: float,
    browser_port: int | None,
    client: bidfax.BidfaxClient | None = None,
    max_concurrent: int = 1,
    makes: dict[str, str] | None = None,
) -> dict[str, tuple[str, str, str]]:
    """Re-fetch bidfax data for each lot. Returns only lots with a result URL.

    `makes` supplies the expected Make for URL validation; when the first
    bidfax hit belongs to a different make the client now returns
    IN_PROGRESS (treated as not-found here) rather than polluting the CSVs.
    """
    real_client = client or bidfax.BrowserBidfaxClient(browser_port=browser_port)
    fetched     = real_client.lookup_many(
        lots, makes=makes, delay=delay, max_concurrent=max_concurrent,
    )
    results: dict[str, tuple[str, str, str]] = {}
    for lot in lots:
        price, vin, url = fetched.get(lot, (bidfax.IN_PROGRESS, "", ""))
        if url:
            results[lot] = (price, vin, url)
            print(f"  [bidfax] {lot} — {price}  VIN:{vin or '—'}  {url}")
        else:
            print(f"  [bidfax] {lot} — not found, SKIPPED")
    return results


# ---------------------------------------------------------------------------
# CSV fix
# ---------------------------------------------------------------------------

def _fix_csv_file(path: Path, results: dict[str, tuple[str, str, str]]) -> int:
    fieldnames, rows = load_csv_dict(path)
    changed = 0
    for row in rows:
        lot = str(row.get(LOT_COL, "")).strip()
        if lot not in results:
            continue
        price, vin, url = results[lot]
        row[PRICE_COL] = price
        if vin:
            row[VIN_COL] = vin
        if url:
            row[LINK_COL] = url
        changed += 1
    if changed:
        save_csv_dict(path, fieldnames, rows)
    return changed


def fix_csvs(directory: Path, results: dict[str, tuple[str, str, str]]) -> int:
    print(f"[*] Scanning {directory} for price CSVs…")
    total = files_touched = 0
    for path in sorted(directory.glob("*.csv")):
        if not PRICE_FILE_PATTERN.match(path.name):
            continue
        changed = _fix_csv_file(path, results)
        if changed:
            files_touched += 1
            total         += changed
            print(f"  [+] {path.name}: {changed} row(s) updated")
    if total:
        print(f"[+] CSVs: {total} row(s) across {files_touched} file(s) updated")
    else:
        print("[*] No matching CSV rows found.")
    return total


# ---------------------------------------------------------------------------
# Workbook fix
# ---------------------------------------------------------------------------

def fix_workbook(workbook_path: Path, results: dict[str, tuple[str, str, str]]) -> int:
    if not workbook_path.exists():
        print(f"[*] Workbook not found: {workbook_path} — skipping")
        return 0

    wb = openpyxl.load_workbook(workbook_path)
    total = 0

    for ws in wb.worksheets:
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
        cols = resolve_columns(headers)
        if cols is None:
            continue

        sheet_updated = 0
        for row in ws.iter_rows(min_row=2):
            lot = str(row[cols.lot - 1].value or "").strip()
            if lot not in results:
                continue
            price, vin, url = results[lot]
            apply_result_to_row(row, cols, price, vin, url)
            sheet_updated += 1

        if sheet_updated:
            print(f"  [+] Sheet {ws.title!r}: {sheet_updated} row(s) updated")
            total += sheet_updated

    if total:
        wb.save(workbook_path)
        print(f"[+] Workbook saved → {workbook_path.name}  ({total} row(s))")
    else:
        print("[*] No matching workbook rows found.")
    return total


# ---------------------------------------------------------------------------
# HTML fix
# ---------------------------------------------------------------------------

def _th_first_text(th) -> str:
    """Return the first text node inside a <th>, ignoring nested spans (sort-icon)."""
    for child in th.children:
        if isinstance(child, NavigableString):
            txt = str(child).strip()
            if txt:
                return txt
    return ""


def _set_link_cell(soup, td, url: str) -> None:
    td.clear()
    if _BIDFAX_DOMAIN in url:
        td["class"] = ["cell-bidfax"]
        label = "Bidfax"
    else:
        td["class"] = ["cell-link"]
        label = "View"
    a = soup.new_tag("a", href=url, target="_blank")
    a.string = label
    td.append(a)


def _update_html_row(soup, tds, headers_idx, result: tuple[str, str, str]) -> None:
    price, vin, url = result
    price_i = headers_idx.get(PRICE_COL, -1)
    vin_i   = headers_idx.get(VIN_COL,   -1)
    link_i  = headers_idx.get(LINK_COL,  -1)

    if 0 <= price_i < len(tds):
        td = tds[price_i]
        td.clear()
        td["class"] = ["cell-price"]
        td.string   = price
    if 0 <= vin_i < len(tds) and vin:
        td = tds[vin_i]
        td.clear()
        td["class"] = ["cell-vin"]
        td.string   = vin
    if 0 <= link_i < len(tds) and url:
        _set_link_cell(soup, tds[link_i], url)


def fix_html(html_path: Path, results: dict[str, tuple[str, str, str]]) -> int:
    if not html_path.exists():
        print(f"[*] HTML not found: {html_path} — skipping")
        return 0
    if not _BS4_OK:
        print("[warn] beautifulsoup4 not installed — skipping HTML update")
        return 0

    soup  = BeautifulSoup(html_path.read_text(encoding="utf-8"), "html.parser")
    total = 0

    for table in soup.select("table.filterable-table"):
        thead = table.find("thead")
        tbody = table.find("tbody")
        if not thead or not tbody:
            continue

        headers = [_th_first_text(th) for th in thead.find_all("th")]
        if LOT_COL not in headers:
            continue

        lot_i       = headers.index(LOT_COL)
        headers_idx = {h: i for i, h in enumerate(headers)}

        for tr in tbody.find_all("tr"):
            if "no-results" in (tr.get("class") or []):
                continue
            tds = tr.find_all("td")
            if lot_i >= len(tds):
                continue
            lot = tds[lot_i].get_text(strip=True)
            if lot not in results:
                continue
            _update_html_row(soup, tds, headers_idx, results[lot])
            total += 1

    if total:
        html_path.write_text(str(soup), encoding="utf-8")
        print(f"[+] HTML: {total} row(s) updated → {html_path.name}")
    else:
        print("[*] No matching HTML rows found.")
    return total


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Re-fetch bidfax data for specific lot numbers and overwrite matching rows in output files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--lots", required=True,
                        help='Comma- or semicolon-separated lot numbers, e.g. "44428368, 44428369"')
    parser.add_argument("--dir",      "-d", default="output",
                        help="Directory containing price CSVs (default: output)")
    parser.add_argument("--workbook", "-w", default=None,
                        help="Excel workbook path (default: <dir>/auction_results.xlsx)")
    parser.add_argument("--html",           default=None,
                        help="HTML report path (default: <dir>/html_report/index.html)")
    parser.add_argument("--delay",          default=2.0, type=float,
                        help="Seconds between bidfax searches (default: 2.0)")
    parser.add_argument("--browser-port",   type=int, default=None,
                        help="Connect to a running Chrome on this port instead of launching one")
    parser.add_argument("--concurrent",     type=int, default=1,
                        help="Parallel bidfax tabs (default: 1 = sequential; experimental)")
    args = parser.parse_args()

    lots = _parse_lots(args.lots)
    if not lots:
        sys.exit("--lots is empty")

    work_dir      = Path(args.dir).resolve()
    workbook_path = Path(args.workbook).resolve() if args.workbook else work_dir / "auction_results.xlsx"
    html_path     = Path(args.html).resolve()     if args.html     else work_dir / "html_report" / "index.html"

    print(f"Lots     : {lots}")
    print(f"Dir      : {work_dir}")
    print(f"Workbook : {workbook_path}")
    print(f"HTML     : {html_path}")
    print(f"Delay    : {args.delay}s\n")

    # Resolve each lot's Make from existing CSVs so bidfax make-validation
    # can reject wrong-vehicle hits (e.g. a Q5 query surfacing a Nissan Leaf).
    makes = find_makes_for_lots(work_dir, lots)
    without_make = [lot for lot in lots if lot not in makes]
    if makes:
        print(f"[*] Resolved Make for {len(makes)}/{len(lots)} lot(s) from CSVs")
    if without_make:
        print(f"[*] No Make found in CSVs for: {', '.join(without_make)} "
              f"— bidfax will accept whatever it returns for these")

    results = lookup_lots(lots, args.delay, args.browser_port,
                          max_concurrent=args.concurrent, makes=makes)

    missing = [l for l in lots if l not in results]
    if missing:
        print(f"\n[*] {len(missing)} lot(s) skipped (not found on bidfax): {', '.join(missing)}")
    if not results:
        print("[!] No lots found on bidfax — nothing to fix.")
        return

    print(f"\n[*] {len(results)} lot(s) resolved — fixing output files")

    print("\n— CSVs —")
    fix_csvs(work_dir, results)

    print("\n— Workbook —")
    fix_workbook(workbook_path, results)

    print("\n— HTML —")
    fix_html(html_path, results)

    print("\n[+] Done.")


if __name__ == "__main__":
    main()
