#!/usr/bin/env python3
"""
Build/update an Excel workbook from auction price CSV files.

Scans the current directory for files matching:
    <auction>_price_<yyyy>_<mm>_<dd>.csv  (iaai or copart)

Files dated today or in the future are skipped. A log file tracks which
files have already been processed so they are never imported twice.

Each unique Make gets its own spreadsheet. The "Last Price" column
(format: "$1,234 | VIN: XXXXXXXX") is split into separate "Price" and
"VIN" columns. Rows where the price is "None" or "In Progress" are skipped.

Usage:
    python build_workbook.py
    python build_workbook.py --workbook my_results.xlsx
    python build_workbook.py --dir /path/to/csvs --log my_log.json
"""

import argparse
import csv
import json
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found. Install with:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

LAST_PRICE_COL  = "Last Price"
PRICE_COL       = "Price"
VIN_COL         = "VIN"

# Matches both  copart_price_…  and  copart_bidfax_…
AUCTION_PATTERN = re.compile(
    r"^(iaai|copart)_(price|bidfax)_(\d{4})_(\d{2})_(\d{2})\.csv$",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# File discovery & log
# ---------------------------------------------------------------------------

def find_pending_files(directory: Path, today: date, processed: set[str]) -> list[Path]:
    """Return price/bidfax CSV files that are older than today and not yet processed."""
    pending = []
    for path in sorted(directory.glob("*.csv")):
        m = AUCTION_PATTERN.match(path.name)
        if not m:
            continue
        try:
            file_date = date(int(m.group(3)), int(m.group(4)), int(m.group(5)))
        except ValueError:
            continue
        if file_date < today and path.name not in processed:
            pending.append(path)
    return pending


def load_log(log_path: Path) -> set[str]:
    if log_path.exists():
        try:
            return set(json.loads(log_path.read_text(encoding="utf-8")))
        except ValueError:
            return set()
    return set()


def save_log(log_path: Path, processed: set[str]) -> None:
    log_path.write_text(json.dumps(sorted(processed), indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# "Last Price" parsing
# ---------------------------------------------------------------------------

def parse_last_price(value: str) -> tuple[str, str]:
    """Split '$1,234 | VIN: ABC123' into ('$1,234', 'ABC123').

    Returns the raw price string as-is (including 'None' or 'In Progress').
    """
    value = value.strip()
    parts = value.split(" | VIN: ", 1)
    price = parts[0].strip()
    vin   = parts[1].strip() if len(parts) > 1 else ""
    return price, vin


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')

def _safe_sheet_name(name: str) -> str:
    return _INVALID_SHEET_CHARS.sub("_", name)[:31]


def _get_or_create_sheet(wb: openpyxl.Workbook, make: str, headers: list[str]):
    """Return the sheet for *make*, creating it with *headers* if needed."""
    title = _safe_sheet_name(make)
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    return ws


def _sheet_headers(ws) -> list[str]:
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    return list(first_row) if first_row else []


def _build_headers(csv_fieldnames: list[str]) -> list[str]:
    """Build output column list.

    - If CSV has 'Last Price' (bidcars format): remove it, insert Price after
      Odometer, append VIN at end.
    - If CSV already has 'Price'/'VIN' (bidfax format): return as-is.
    """
    if LAST_PRICE_COL in csv_fieldnames:
        cols = [c for c in csv_fieldnames if c != LAST_PRICE_COL]
        insert_at = cols.index("Odometer") + 1 if "Odometer" in cols else len(cols)
        cols.insert(insert_at, PRICE_COL)
        if VIN_COL not in cols:
            cols.append(VIN_COL)
        return cols
    return list(csv_fieldnames)


# ---------------------------------------------------------------------------
# CSV processing
# ---------------------------------------------------------------------------

def _cell_value(col: str, row: dict, price: str, vin: str) -> str:
    if col == PRICE_COL:
        return price
    if col == VIN_COL:
        return vin
    if col == "Link":
        url = row.get("Link", "").strip()
        return f'=HYPERLINK("{url}")' if url else ""
    return row.get(col, "")


def _extract_price_vin(row: dict, fieldnames: list[str]) -> tuple[str, str]:
    """Return (price, vin) from a row regardless of CSV format."""
    if LAST_PRICE_COL in fieldnames:
        return parse_last_price(row.get(LAST_PRICE_COL, ""))
    price = row.get(PRICE_COL, "").strip()
    vin   = row.get(VIN_COL, "").strip()
    return price, vin


def process_csv(csv_path: Path, wb: openpyxl.Workbook) -> int:
    """Import one CSV into the workbook. Returns count of rows added."""
    added = 0
    sheet_hdrs: dict[str, list[str]] = {}

    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            print(f"    [warn] No headers found — skipping {csv_path.name}")
            return 0

        fieldnames = list(reader.fieldnames)
        headers    = _build_headers(fieldnames)

        for row in reader:
            price, vin = _extract_price_vin(row, fieldnames)
            make  = (row.get("Make") or "UNKNOWN").strip().upper()
            ws    = _get_or_create_sheet(wb, make, headers)
            title = ws.title
            if title not in sheet_hdrs:
                sheet_hdrs[title] = _sheet_headers(ws)
            ws.append([_cell_value(col, row, price, vin) for col in sheet_hdrs[title]])
            added += 1

    return added


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build/update an Excel workbook from auction price CSV files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--workbook", "-w", default="auction_results.xlsx",
                        help="Excel workbook to create/update (default: auction_results.xlsx)")
    parser.add_argument("--log",      "-l", default="processed_files.json",
                        help="Log of already-processed files (default: processed_files.json)")
    parser.add_argument("--dir",      "-d", default=".",
                        help="Directory to scan for CSV files (default: current dir)")
    args = parser.parse_args()

    work_dir      = Path(args.dir).resolve()
    workbook_path = work_dir / args.workbook
    log_path      = work_dir / args.log
    today         = date.today()

    processed = load_log(log_path)
    pending   = find_pending_files(work_dir, today, processed)

    all_count = len(find_pending_files(work_dir, today, set()))
    print(f"[*] Matching files : {all_count}")
    print(f"[*] Already processed : {len(processed)}")
    print(f"[*] To process now    : {len(pending)}")

    if not pending:
        print("[+] Nothing new to process.")
        return

    if workbook_path.exists():
        print(f"[*] Opening:  {workbook_path.name}")
        wb = openpyxl.load_workbook(workbook_path)
    else:
        print(f"[*] Creating: {workbook_path.name}")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)          # drop the default blank sheet

    total_added = 0
    for csv_path in pending:
        print(f"\n  [*] {csv_path.name}")
        added = process_csv(csv_path, wb)
        print(f"      added: {added}")
        total_added += added
        processed.add(csv_path.name)

    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws.append(["No data was imported."])
        print("[warn] No sheets created — adding placeholder sheet.")

    wb.save(workbook_path)
    save_log(log_path, processed)

    print(f"\n[+] Workbook saved : {workbook_path}")
    print(f"[+] Log updated    : {log_path}")
    print(f"[+] Total added    : {total_added}")


if __name__ == "__main__":
    main()
